#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
单元测试：detection_logger.py

测试日志记录和数据序列化
"""

import glob
import json
import tempfile
import unittest
from datetime import date
from pathlib import Path
from unittest import mock

from db_manager import PatentsDB
from detection_logger import DetectionLogger, DetectionRecord


class TestDetectionRecord(unittest.TestCase):
    """DetectionRecord 数据结构测试"""

    def test_record_creation(self):
        """创建基础记录"""
        record = DetectionRecord(
            application_no='CN201880002233',
            status_code=200,
            response_summary='success'
        )
        self.assertEqual(record.application_no, 'CN201880002233')
        self.assertEqual(record.status_code, 200)
        self.assertEqual(record.response_summary, 'success')

    def test_record_with_patent_fields(self):
        """创建包含专利字段的记录"""
        record = DetectionRecord(
            application_no='CN201880002233',
            status_code=200,
            zhuanlimc='一种新型装置',
            shenqingrxm='申请人名字',
            anjianywzt='驳回等复审请求'
        )
        self.assertEqual(record.zhuanlimc, '一种新型装置')
        self.assertEqual(record.shenqingrxm, '申请人名字')
        self.assertEqual(record.anjianywzt, '驳回等复审请求')

    def test_record_with_fwxx(self):
        """创建包含发文信息的记录"""
        fwxx_list = [
            {'date': '2023-01-01', 'type': '驳回'},
            {'date': '2023-02-01', 'type': '再审'}
        ]
        record = DetectionRecord(
            application_no='CN201880002233',
            status_code=200,
            fwxx_list=fwxx_list
        )
        self.assertEqual(len(record.fwxx_list), 2)
        self.assertEqual(record.fwxx_list[0]['type'], '驳回')

    def test_record_serializes_independent_fwxx_collection_time(self):
        record = DetectionRecord(
            application_no='CN201880002233',
            fwxx_list=[],
            fwxx_collected_at='2026-09-06T12:00:00Z',
        )
        self.assertEqual(
            record.to_dict()['fwxx_collected_at'],
            '2026-09-06T12:00:00Z',
        )

    def test_record_with_complete_fee_snapshot(self):
        payable = [{'yingjiaoffyzlmc': '实用新型专利第6年年费'}]
        late_schedule = [{'zhinajjfsj': '2026年01月06日到2026年02月03日'}]
        record = DetectionRecord(
            application_no='2020228959227',
            payable_fee_records=payable,
            late_fee_schedule_records=late_schedule,
            fee_snapshot_at='2026-07-18T00:00:00Z',
        )
        serialized = record.to_dict()
        self.assertEqual(serialized['payable_fee_records'], payable)
        self.assertEqual(serialized['late_fee_schedule_records'], late_schedule)
        self.assertEqual(serialized['fee_snapshot_at'], '2026-07-18T00:00:00Z')

    def test_record_to_dict(self):
        """序列化为字典"""
        record = DetectionRecord(
            application_no='CN201880002233',
            status_code=200,
            zhuanlimc='测试专利'
        )
        record_dict = record.to_dict()
        self.assertIsInstance(record_dict, dict)
        self.assertEqual(record_dict['application_no'], 'CN201880002233')
        self.assertEqual(record_dict['status_code'], 200)
        self.assertEqual(record_dict['zhuanlimc'], '测试专利')

    def test_failed_record(self):
        """失败记录"""
        record = DetectionRecord(
            application_no='CN201880002233',
            status_code=0,
            response_summary='MITM timeout',
            error_message='8s timeout'
        )
        self.assertEqual(record.status_code, 0)
        self.assertEqual(record.error_message, '8s timeout')

    def test_record_with_response_time(self):
        """包含响应时间的记录"""
        record = DetectionRecord(
            application_no='CN201880002233',
            status_code=200,
            response_time_ms=3456.78
        )
        self.assertEqual(record.response_time_ms, 3456.78)

    def test_record_timestamp(self):
        """时间戳"""
        record = DetectionRecord(
            application_no='CN201880002233',
            status_code=200
        )
        self.assertIsNotNone(record.timestamp)
        # timestamp 应该是 ISO 格式
        self.assertIn('T', record.timestamp)

    def test_empty_patent_fields(self):
        """空专利字段"""
        record = DetectionRecord(
            application_no='CN201880002233',
            status_code=0  # 失败
        )
        self.assertIsNone(record.zhuanlimc)
        self.assertIsNone(record.shenqingrxm)
        self.assertIsNone(record.anjianywzt)


class TestDetectionRecordSerialization(unittest.TestCase):
    """序列化和反序列化测试"""

    def test_json_serializable(self):
        """记录可被序列化为 JSON"""
        record = DetectionRecord(
            application_no='CN201880002233',
            status_code=200,
            zhuanlimc='测试'
        )
        record_dict = record.to_dict()
        json_str = json.dumps(record_dict, ensure_ascii=False)
        self.assertIsInstance(json_str, str)

    def test_json_deserialization(self):
        """从 JSON 反序列化"""
        record_dict = {
            'application_no': 'CN201880002233',
            'status_code': 200,
            'zhuanlimc': '测试',
            'timestamp': '2026-05-11T12:00:00'
        }
        json_str = json.dumps(record_dict)
        loaded_dict = json.loads(json_str)

        record = DetectionRecord(
            application_no=loaded_dict['application_no'],
            status_code=loaded_dict['status_code'],
            zhuanlimc=loaded_dict.get('zhuanlimc')
        )
        self.assertEqual(record.application_no, 'CN201880002233')
        self.assertEqual(record.zhuanlimc, '测试')

    def test_unicode_support(self):
        """支持中文和 Unicode"""
        record = DetectionRecord(
            application_no='CN201880002233',
            status_code=200,
            zhuanlimc='一种新型装置',
            shenqingrxm='张三李四',
            response_summary='成功采集'
        )
        record_dict = record.to_dict()
        json_str = json.dumps(record_dict, ensure_ascii=False)
        loaded = json.loads(json_str)

        self.assertEqual(loaded['zhuanlimc'], '一种新型装置')
        self.assertEqual(loaded['shenqingrxm'], '张三李四')


def _logger_in(tmpdir: str, name: str = 'log.jsonl') -> DetectionLogger:
    """在临时目录中构造 DetectionLogger（DB 也落在临时目录，不碰仓库数据）"""
    db_path = Path(tmpdir) / 'patents.db'
    with mock.patch('detection_logger.PATENTS_DB_FILE', db_path):
        return DetectionLogger(str(Path(tmpdir) / name))


class TestDetectionLoggerWritePath(unittest.TestCase):
    """DB + JSONL 双写路径"""

    def test_add_record_appends_jsonl_and_db(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            logger = _logger_in(tmpdir)
            logger.add_record(DetectionRecord(application_no='2023000000001', status_code=200))
            logger.add_record(DetectionRecord(application_no='2023000000002', status_code=0))
            lines = Path(logger.log_file).read_text(encoding='utf-8').splitlines()
            self.assertEqual(len(lines), 2)
            self.assertEqual(logger._db.count(), 2)

    def test_auto_backup_fires_on_interval_and_resets_counter(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            logger = _logger_in(tmpdir)
            logger.add_record(DetectionRecord(application_no='2023000000001', status_code=200))
            backup_pattern = str(Path(tmpdir) / 'log_backup_*.db')

            logger._auto_backup(written=500)
            self.assertEqual(len(glob.glob(backup_pattern)), 1)
            self.assertEqual(logger._writes_since_backup, 0)

            logger._auto_backup(written=1)
            self.assertEqual(len(glob.glob(backup_pattern)), 1)
            self.assertEqual(logger._writes_since_backup, 1)

    def test_auto_backup_includes_upsert_updates_absent_from_jsonl(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            logger = _logger_in(tmpdir)
            logger.add_record(DetectionRecord(
                application_no='2023000000001', status_code=0, zhuanlimc='old title',
            ))
            logger._writes_since_backup = 499

            logger.upsert_record(DetectionRecord(
                application_no='2023000000001', status_code=200, zhuanlimc='updated title',
            ))

            self.assertEqual(json.loads(Path(logger.log_file).read_text())['status_code'], 0)
            backup_paths = list(Path(tmpdir).glob('log_backup_*.db'))
            self.assertEqual(len(backup_paths), 1)
            restored = PatentsDB(backup_paths[0]).get_record('2023000000001')
            self.assertEqual(restored['status_code'], 200)
            self.assertEqual(restored['zhuanlimc'], 'updated title')

    def test_failed_auto_backup_keeps_counter_for_retry(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            logger = _logger_in(tmpdir)
            with mock.patch.object(logger._db, 'backup_to', side_effect=OSError('disk unavailable')):
                with self.assertRaises(OSError):
                    logger._auto_backup(written=500)

            self.assertEqual(logger._writes_since_backup, 500)
            self.assertEqual(list(Path(tmpdir).glob('log_backup_*.db')), [])
            logger._auto_backup(written=0)
            self.assertEqual(logger._writes_since_backup, 0)
            self.assertEqual(len(list(Path(tmpdir).glob('log_backup_*.db'))), 1)

    def test_database_backup_retention_preserves_five_and_historical_jsonl(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            logger = _logger_in(tmpdir)
            root = Path(tmpdir)
            historical_backup = root / 'log_backup_20200101_000000.jsonl'
            historical_backup.write_text('historical backup\n', encoding='utf-8')
            for sequence in range(7):
                logger._db.backup_to(root / f'log_backup_20260101_00000{sequence}.db')

            logger._prune_backups()

            self.assertEqual(
                sorted(backup.name for backup in root.glob('log_backup_*.db')),
                [f'log_backup_20260101_00000{sequence}.db' for sequence in range(2, 7)],
            )
            self.assertEqual(historical_backup.read_text(encoding='utf-8'), 'historical backup\n')

    def test_add_records_matches_looped_add_record(self):
        records = [
            DetectionRecord(application_no='2023000000001', status_code=200, zhuanlimc='专利一'),
            DetectionRecord(application_no='2023000000002', status_code=0, error_message='超时'),
            DetectionRecord(application_no='2023000000003', status_code=200,
                            fwxx_list=[{'fawenmc': '驳回决定'}]),
        ]
        with tempfile.TemporaryDirectory() as loop_dir, \
                tempfile.TemporaryDirectory() as batch_dir:
            looped = _logger_in(loop_dir)
            for r in records:
                looped.add_record(r)

            batched = _logger_in(batch_dir)
            self.assertEqual(batched.add_records(records), 3)

            def without_write_stamp(rows: list) -> list:
                # updated_at 由 DB 在写入时刻生成，两个 logger 必然不同，剔除后比较
                return [{k: v for k, v in row.items() if k != 'updated_at'} for row in rows]

            self.assertEqual(
                without_write_stamp(batched._db.get_all_records()),
                without_write_stamp(looped._db.get_all_records()),
            )
            self.assertEqual(
                Path(batched.log_file).read_text(encoding='utf-8'),
                Path(looped.log_file).read_text(encoding='utf-8'),
            )


class TestFeeExcelExport(unittest.TestCase):
    def test_exports_all_fee_sheets_with_text_identifiers(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            logger = _logger_in(tmpdir)
            logger.add_record(DetectionRecord(
                application_no='2026102909420',
                status_code=200,
                payable_fee_records=[{
                    'yingjiaoffyzlmc': '发明专利第6年年费',
                    'yingjiaoje': 1200,
                    'jiaofeijzr': '2026-06-03',
                    'yingjiaoffyzt': '未缴',
                }],
                late_fee_schedule_records=[{
                    'zhinajjfsj': '2026年01月06日到2026年02月03日',
                    'zhinajdqnfje': 1200,
                    'zhinajyjznje': 60,
                    'zhinajzj': 1260,
                }],
                paid_fee_records=[{
                    'yijiaofjfzlmc': '发明专利申请实质审查费',
                    'yijiaofjfje': 375,
                    'yijiaofjfrq': '2026-03-11',
                    'yijiaofjfrxm': '某专利代理事务所',
                    'yijiaofpjdm': '00010125',
                    'yijiaofpjhm': '0026303067',
                }],
                fee_receipt_dispatch_records=[{
                    'shoujufwfyzlmc': '发明专利申请实质审查费',
                    'shoujufwjfje': 375,
                    'shoujufwjfrxm': '某专利代理事务所',
                    'shoujufwjfsj': '2026-03-11',
                    'shoujufwsjh': '0026303067',
                    'shoujufwsjtt': '某专利代理事务所',
                    'shoujufwyjdz': '',
                    'shoujufwtkrq': '',
                    'shoujufwsfjc': '',
                    'shoujufwfwrq': '',
                    'shoujufwghhm': '',
                    'shoujufwtkhcrq': '',
                }],
                fwxx_collected_at='2026-07-18T07:30:00Z',
                fee_snapshot_at='2026-07-18T00:00:00Z',
            ))
            excel_path = Path(tmpdir) / 'fees.xlsx'

            self.assertTrue(logger.export_to_excel(
                str(excel_path),
                fee_analysis_date=date(2026, 2, 3),
            ))

            workbook = load_workbook(excel_path)
            self.assertIn('应缴费信息', workbook.sheetnames)
            self.assertIn('应缴滞纳金信息', workbook.sheetnames)
            self.assertIn('待缴费分析', workbook.sheetnames)
            self.assertIn('当前滞纳金', workbook.sheetnames)
            self.assertIn('已缴费信息', workbook.sheetnames)
            self.assertIn('收据发文信息', workbook.sheetnames)
            payable_sheet = workbook['应缴费信息']
            late_fee_sheet = workbook['应缴滞纳金信息']
            payable_analysis_sheet = workbook['待缴费分析']
            late_fee_analysis_sheet = workbook['当前滞纳金']
            paid_sheet = workbook['已缴费信息']
            receipt_sheet = workbook['收据发文信息']
            patent_sheet = workbook['专利主信息']
            patent_headers = [cell.value for cell in patent_sheet[1]]
            fwxx_time_column = patent_headers.index('发文采集时间') + 1
            self.assertEqual(
                patent_sheet.cell(row=2, column=fwxx_time_column).value,
                '2026-07-18T07:30:00Z',
            )
            self.assertEqual(payable_sheet['A2'].value, '2026102909420')
            self.assertEqual(payable_sheet['A2'].data_type, 's')
            self.assertEqual(payable_sheet['A2'].number_format, '@')
            self.assertEqual(payable_sheet['B2'].value, '发明专利第6年年费')
            self.assertEqual(late_fee_sheet['B2'].value, '2026年01月06日到2026年02月03日')
            self.assertEqual(late_fee_sheet['A2'].number_format, '@')
            self.assertEqual(payable_analysis_sheet['B2'].value, '2026102909420')
            self.assertEqual(payable_analysis_sheet['G2'].value, '发明专利第6年年费')
            self.assertEqual(payable_analysis_sheet['L2'].value, '未来')
            self.assertEqual(late_fee_analysis_sheet['F2'].value, '当前适用')
            self.assertEqual(late_fee_analysis_sheet['G2'].value, '2026年01月06日到2026年02月03日')
            self.assertEqual(late_fee_analysis_sheet['K2'].value, 60)
            self.assertEqual(late_fee_analysis_sheet['L2'].value, 1260)
            self.assertEqual(paid_sheet['F2'].value, '00010125')
            self.assertEqual(paid_sheet['F2'].data_type, 's')
            self.assertEqual(paid_sheet['F2'].number_format, '@')
            self.assertEqual(paid_sheet['G2'].value, '0026303067')
            self.assertEqual(paid_sheet['G2'].data_type, 's')
            self.assertEqual(paid_sheet['G2'].number_format, '@')
            self.assertEqual(receipt_sheet['F2'].value, '0026303067')
            self.assertEqual(receipt_sheet['F2'].data_type, 's')
            self.assertEqual(receipt_sheet['F2'].number_format, '@')
            self.assertEqual(paid_sheet.freeze_panes, 'A2')
            self.assertGreaterEqual(paid_sheet.column_dimensions['E'].width, 40)
            self.assertEqual(receipt_sheet.max_column, 13)
            workbook.close()


if __name__ == '__main__':
    unittest.main()
