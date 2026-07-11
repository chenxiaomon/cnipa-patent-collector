#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CNIPA 采集系统本地可视化控制台。

设计目标：
- 不引入 Flask/FastAPI/Node 等新依赖，直接用 Python 标准库启动本地 Web UI。
- 只通过白名单命令调用现有脚本，避免把命令行参数暴露成任意 shell。
- 后台任务统一收集日志，方便从浏览器观察 MITM、采集、策略生成等流程。
"""

from __future__ import annotations

import argparse
import io
import json
import mimetypes
import os
import re
import socket
import subprocess
import sys
import threading
import time
import uuid
from collections import deque
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

from settings import (
    BASE_DIR,
    COMPANY_META_FILE,
    CONFIG_FILE,
    DATA_DIR,
    DETECTION_LOG_FILE,
    DETECTION_LOG_JSONL_FILE,
    MITM_HOST,
    MITM_PORT,
    PATENTS_DB_FILE,
    PATENTS_EXCEL_FILE,
    RETRY_FAILED_FILE,
    SEARCH_LIST_FILE,
    CNIPA_LOGIN_WAIT_SECONDS,
)
from db_manager import PatentsDB
from cache_utils import normalize_app_no, parse_app_no_list, parse_timestamp
from machine_identity import MASTER_ROLE, read_machine_role
from collection_health import read_alert_status
from operator_api_token import api_token_matches, ensure_api_token

_patents_db = PatentsDB(PATENTS_DB_FILE)


APP_NAME = "CNIPA 采集控制台"
SERVER_VERSION = "CNIPADashboard/0.2"
MAX_LOG_LINES = 1600
# 传给采集子进程的登录等待时间，与 settings.CNIPA_LOGIN_WAIT_SECONDS 保持一致
DEFAULT_LOGIN_WAIT_SECONDS = str(int(CNIPA_LOGIN_WAIT_SECONDS))
MAX_BODY_BYTES = 1 * 1024 * 1024   # 1 MB：防止超大请求体撑爆内存
MAX_REQUEST_APP_NOS = 500           # 单次提交申请号上限
MAX_NOTE_LEN = 500                  # 备注字段长度上限
_SAFE_ID_RE = re.compile(r'^[0-9a-f\-]{8,36}$')
DESKTOP_BROWSER_ACTIONS = {
    "main_full",
    "main_test",
    "main_update_dynamic",
    "collect_fwxx",
    "collect_fwxx_app",
    "phase0_browser",
    "public_browser",
    "public_auto_paginate",
    "retry_failed_run_batch",
    "strategy_collect",
}


DOWNLOADS = {
    "excel": PATENTS_EXCEL_FILE,
    "jsonl": DETECTION_LOG_JSONL_FILE,
    "json": DETECTION_LOG_FILE,
    "dynamic": DATA_DIR / "update_list_dynamic.txt",
    "retry": DATA_DIR / "retry_dynamic.txt",
}


def _parse_path_segment(path: str, index: int) -> str:
    """提取 URL 路径段，验证格式（仅小写十六进制和连字符）。"""
    parts = path.split("/")
    if len(parts) <= index:
        raise ValueError("路径格式不正确")
    segment = parts[index]
    if not segment or not _SAFE_ID_RE.match(segment):
        raise ValueError(f"ID 格式不正确: {segment!r}")
    return segment


def _write_text_atomic(path: Path, text: str) -> None:
    """原子写入文本文件（.tmp + replace 保证写入完整性）。"""
    tmp = path.with_suffix(path.suffix + '.tmp')
    tmp.write_text(text, encoding='utf-8')
    tmp.replace(path)


def _parse_env_file(env_path: Path) -> dict[str, str]:
    """解析 .env 文件为 key→value 字典，跳过空行和注释行。"""
    if not env_path.exists():
        return {}
    pairs: dict[str, str] = {}
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith('#') or '=' not in line:
            continue
        k, _, v = line.partition('=')
        pairs[k.strip()] = v.strip()
    return pairs


def resolve_task_python() -> str:
    """后台任务优先使用项目虚拟环境，避免系统 Python 缺依赖。"""
    override = os.getenv("DASHBOARD_TASK_PYTHON")
    if override:
        return override

    for candidate in (
        BASE_DIR / ".venv" / "bin" / "python",
        BASE_DIR / "venv" / "bin" / "python",
        BASE_DIR / ".venv" / "Scripts" / "python.exe",
        BASE_DIR / "venv" / "Scripts" / "python.exe",
    ):
        if candidate.exists():
            return str(candidate)

    return sys.executable


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def iso_now() -> str:
    return utc_now().isoformat().replace("+00:00", "Z")


def safe_read_text(path: Path, default: str = "") -> str:
    try:
        return path.read_text(encoding="utf-8")
    except FileNotFoundError:
        return default


def safe_json_load(path: Path, default: Any) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return default


def count_lines(path: Path) -> int:
    try:
        with path.open("r", encoding="utf-8") as handle:
            return sum(1 for line in handle if line.strip())
    except FileNotFoundError:
        return 0


def file_info(path: Path) -> dict[str, Any]:
    exists = path.exists()
    stat = path.stat() if exists else None
    return {
        "exists": exists,
        "name": path.name,
        "path": str(path.relative_to(BASE_DIR) if path.is_relative_to(BASE_DIR) else path),
        "size": stat.st_size if stat else 0,
        "mtime": datetime.fromtimestamp(stat.st_mtime, timezone.utc).isoformat().replace("+00:00", "Z")
        if stat
        else None,
    }


def port_open(host: str, port: int, timeout: float = 0.2) -> bool:
    try:
        with socket.create_connection((host, port), timeout=timeout):
            return True
    except OSError:
        return False


@dataclass
class Job:
    id: str
    action: str
    title: str
    command: list[str]
    env_overrides: dict[str, str] = field(default_factory=dict)
    started_at: str = field(default_factory=iso_now)
    finished_at: str | None = None
    status: str = "running"
    returncode: int | None = None
    lines: deque[str] = field(default_factory=lambda: deque(maxlen=MAX_LOG_LINES))
    process: subprocess.Popen[str] | None = None

    def append(self, line: str) -> None:
        self.lines.append(line.rstrip("\n"))

    def to_dict(self, include_logs: bool = False) -> dict[str, Any]:
        lines_list = list(self.lines)
        waiting = (
            self.status == "running"
            and any("[WAITING_FOR_LOGIN]" in ln for ln in lines_list)
            and not any(
                "收到登录完成信号" in ln or "秒超时，继续执行" in ln
                for ln in lines_list
            )
        )
        data = {
            "id": self.id,
            "action": self.action,
            "title": self.title,
            "command": printable_command(self.command),
            "started_at": self.started_at,
            "finished_at": self.finished_at,
            "status": self.status,
            "returncode": self.returncode,
            "log_count": len(self.lines),
            "waiting_for_login": waiting,
        }
        if include_logs:
            data["logs"] = lines_list
        return data


def printable_command(command: list[str]) -> str:
    return " ".join(command)


class JobManager:
    def __init__(self) -> None:
        self._jobs: dict[str, Job] = {}
        self._lock = threading.Lock()

    def list_jobs(self) -> list[dict[str, Any]]:
        with self._lock:
            jobs = sorted(self._jobs.values(), key=lambda item: item.started_at, reverse=True)
            return [job.to_dict() for job in jobs[:40]]

    def get_job(self, job_id: str) -> Job | None:
        with self._lock:
            return self._jobs.get(job_id)

    def start(self, action: str, params: dict[str, Any]) -> Job:
        spec = build_job_spec(action, params)
        if self._already_running(spec["action"]):
            raise ValueError(f"{spec['title']} 已在运行")

        job = Job(
            id=uuid.uuid4().hex[:10],
            action=spec["action"],
            title=spec["title"],
            command=spec["command"],
            env_overrides=spec.get("env", {}),
        )
        env = os.environ.copy()
        env.update(job.env_overrides)
        env.setdefault("PYTHONUNBUFFERED", "1")
        env.setdefault("PYTHONIOENCODING", "utf-8")

        requires_desktop = job.action in DESKTOP_BROWSER_ACTIONS
        extra_kwargs = {}
        if sys.platform == 'win32' and not requires_desktop:
            extra_kwargs['creationflags'] = subprocess.CREATE_NO_WINDOW

        process = subprocess.Popen(
            job.command,
            cwd=str(BASE_DIR),
            env=env,
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding='utf-8',
            bufsize=1,
            **extra_kwargs,
        )
        job.process = process
        if process.stdin:
            process.stdin.close()
        job.append(f"$ {printable_command(job.command)}")
        if requires_desktop:
            job.append("[dashboard] 此任务会在运行 Dashboard 的机器上启动/控制浏览器；远程访问网页不会在客户端电脑弹出浏览器。")

        with self._lock:
            self._jobs[job.id] = job

        thread = threading.Thread(target=self._watch_process, args=(job,), daemon=True)
        thread.start()
        return job

    def stop(self, job_id: str) -> bool:
        with self._lock:
            job = self._jobs.get(job_id)
            if not job or not job.process:
                return False
            if job.process.poll() is not None:
                return False
            job.status = "stopping"
        job.append("[dashboard] 正在停止任务...")
        job.process.terminate()
        threading.Thread(target=self._force_kill_later, args=(job,), daemon=True).start()
        return True

    def _already_running(self, action: str) -> bool:
        singleton_actions = {"mitm_proxy", "public_mitm_proxy", "phase0_browser", "public_browser"}
        if action not in singleton_actions:
            return False
        with self._lock:
            for job in self._jobs.values():
                if job.action == action and job.status in {"running", "stopping"}:
                    return True
        return False

    def _watch_process(self, job: Job) -> None:
        assert job.process is not None
        assert job.process.stdout is not None
        try:
            for line in job.process.stdout:
                job.append(line)
        finally:
            job.process.stdout.close()
            returncode = job.process.wait()
            with self._lock:
                job.returncode = returncode
                job.finished_at = iso_now()
                if job.status == "stopping":
                    job.status = "stopped"
                elif returncode == 0:
                    job.status = "finished"
                else:
                    job.status = "failed"
            job.append(f"[dashboard] 任务结束，退出码: {returncode}")

    def _force_kill_later(self, job: Job) -> None:
        time.sleep(5)
        if job.process and job.process.poll() is None:
            job.append("[dashboard] 任务未正常退出，强制结束")
            job.process.kill()


def positive_int(value: Any, default: int | None = None, minimum: int = 1, maximum: int = 100000) -> int | None:
    if value in (None, ""):
        return default
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return max(minimum, min(maximum, parsed))


def relative_data_file(value: str | None, default: str) -> str:
    text = (value or default).strip()
    candidate = (BASE_DIR / text).resolve()
    data_root = DATA_DIR.resolve()
    if candidate == data_root or data_root not in candidate.parents:
        return default
    return str(candidate.relative_to(BASE_DIR))


def build_job_spec(action: str, params: dict[str, Any]) -> dict[str, Any]:
    py = resolve_task_python()
    action = action.strip()

    if action == "mitm_proxy":
        return {"action": action, "title": "主 MITM 代理", "command": [py, "-u", "start_mitm_proxy.py"]}
    if action == "public_mitm_proxy":
        return {"action": action, "title": "公开查询 MITM 代理", "command": [py, "-u", "start_mitm_public_search.py"]}
    if action == "main_full":
        return {
            "action": action, "title": "看门狗主流程采集",
            "command": [py, "-u", "collection_watchdog.py"],
            "env": {"USE_MITM_PROXY": "true", "CNIPA_LOGIN_WAIT_SECONDS": DEFAULT_LOGIN_WAIT_SECONDS},
        }
    if action == "main_test":
        count = positive_int(params.get("count"), default=5, maximum=10000)
        return {
            "action": action, "title": f"强制测试前 {count} 条",
            "command": [py, "-u", "main_automation.py", "--update-list", "data/search_list.txt", "--test", str(count)],
            "env": {"USE_MITM_PROXY": "true", "CNIPA_LOGIN_WAIT_SECONDS": DEFAULT_LOGIN_WAIT_SECONDS},
        }
    if action == "main_update_dynamic":
        update_file = relative_data_file(params.get("file"), "data/update_list_dynamic.txt")
        command = [py, "-u", "main_automation.py", "--update-list", update_file]
        count = positive_int(params.get("count"), default=None, maximum=10000)
        if count:
            command.extend(["--test", str(count)])
        return {
            "action": action, "title": f"按清单更新 {Path(update_file).name}",
            "command": command,
            "env": {"USE_MITM_PROXY": "true", "CNIPA_LOGIN_WAIT_SECONDS": DEFAULT_LOGIN_WAIT_SECONDS},
        }
    if action == "collect_fwxx":
        command = [py, "-u", "collect_fwxx.py"]
        count = positive_int(params.get("count"), default=None, maximum=10000)
        if count:
            command.extend(["--test", str(count)])
        return {
            "action": action, "title": "补采发文信息", "command": command,
            "env": {"USE_MITM_PROXY": "true", "CNIPA_LOGIN_WAIT_SECONDS": DEFAULT_LOGIN_WAIT_SECONDS},
        }
    if action == "collect_fwxx_app":
        app_no = normalize_app_no(params.get("app_no"))
        if not app_no:
            raise ValueError("请输入申请号")
        return {
            "action": action, "title": f"补采发文 {app_no}",
            "command": [py, "-u", "collect_fwxx.py", "--app", app_no],
            "env": {"USE_MITM_PROXY": "true", "CNIPA_LOGIN_WAIT_SECONDS": DEFAULT_LOGIN_WAIT_SECONDS},
        }
    if action == "strategy_generate":
        command = [py, "-u", "update_by_strategy.py", "generate"]
        freq = positive_int(params.get("frequency"), default=None, maximum=3650)
        if freq:
            command.append(str(freq))
        return {"action": action, "title": "生成状态检查清单", "command": command}
    if action == "strategy_collect":
        command = [py, "-u", "run_strategy_update.py"]
        freq = positive_int(params.get("frequency"), default=None, maximum=3650)
        if freq:
            command.append(str(freq))
        count = positive_int(params.get("count"), default=None, maximum=10000)
        if count:
            command.extend(["--test", str(count)])
        title = f"采集 {freq} 天策略组" if freq else "采集全部策略组"
        return {
            "action": action,
            "title": title,
            "command": command,
            "env": {"USE_MITM_PROXY": "true", "CNIPA_LOGIN_WAIT_SECONDS": DEFAULT_LOGIN_WAIT_SECONDS},
        }
    if action == "strategy_status":
        command = [py, "-u", "update_by_strategy.py", "status"]
        freq = positive_int(params.get("frequency"), default=None, maximum=3650)
        if freq:
            command.append(str(freq))
        return {"action": action, "title": "查看策略状态", "command": command}
    if action == "strategy_check":
        app_no = normalize_app_no(params.get("app_no"))
        if not app_no:
            raise ValueError("请输入申请号")
        return {
            "action": action, "title": f"检查申请号 {app_no}",
            "command": [py, "-u", "update_by_strategy.py", "check", app_no],
        }
    if action in {"strategy_prepare", "strategy_diff", "strategy_report", "strategy_validate", "strategy_stats"}:
        cmd_map = {
            "strategy_prepare": "prepare", "strategy_diff": "diff",
            "strategy_report": "report", "strategy_validate": "validate", "strategy_stats": "stats",
        }
        title_map = {
            "strategy_prepare": "保存采集前快照", "strategy_diff": "查看状态变化",
            "strategy_report": "生成详细报告", "strategy_validate": "验证策略计数", "strategy_stats": "策略统计",
        }
        return {
            "action": action, "title": title_map[action],
            "command": [py, "-u", "update_by_strategy.py", cmd_map[action]],
        }
    if action == "export_excel":
        return {
            "action": action, "title": "导出 Excel",
            "command": [py, "-u", "-c", "from detection_logger import DetectionLogger; DetectionLogger().export_to_excel()"],
        }
    if action == "export_json":
        return {
            "action": action, "title": "导出 JSON",
            "command": [py, "-u", "-c", "from detection_logger import DetectionLogger; DetectionLogger().export_to_json()"],
        }
    if action == "phase0_browser":
        return {"action": action, "title": "Phase 0 浏览器", "command": [py, "-u", "start_browser_for_phase0.py"]}
    if action == "import_cache":
        return {"action": action, "title": "导入 MITM 缓存", "command": [py, "-u", "import_from_cache.py"]}
    if action == "import_public_search":
        return {"action": action, "title": "导入公开查询结果", "command": [py, "-u", "import_public_search.py"]}
    if action == "public_browser":
        return {"action": action, "title": "公开查询浏览器", "command": [py, "-u", "launch_browser_with_proxy.py"]}
    if action == "public_auto_paginate":
        delay = params.get("delay", 1.5)
        try:
            delay_text = str(max(0.2, min(30.0, float(delay))))
        except (TypeError, ValueError):
            delay_text = "1.5"
        max_pages = positive_int(params.get("max_pages"), default=50, maximum=10000)
        return {
            "action": action, "title": f"公开查询自动翻页 {max_pages} 页",
            "command": [py, "-u", "auto_paginate.py", "--delay", delay_text, "--max-pages", str(max_pages)],
        }
    if action == "public_export":
        return {"action": action, "title": "导出公开查询结果", "command": [py, "-u", "export_public_search.py"]}
    # ── 数据管理类（新增）──────────────────────────────────────────
    if action == "retry_failed":
        return {"action": action, "title": "生成失败重试清单", "command": [py, "-u", "retry_failed.py", "--write-list"]}
    if action == "retry_failed_batch":
        batch_size = positive_int(params.get("batch_size"), default=200, maximum=10000)
        batch_file = relative_data_file(params.get("batch_file"), "data/retry_batch_001.txt")
        return {
            "action": action,
            "title": f"生成失败重试批次 {batch_size} 条",
            "command": [py, "-u", "retry_failed.py", "--batch-size", str(batch_size), "--batch-file", batch_file],
        }
    if action == "retry_failed_run_batch":
        batch_file = relative_data_file(params.get("batch_file"), "data/retry_batch_001.txt")
        timeout = positive_int(params.get("timeout"), default=12, maximum=120)
        return {
            "action": action,
            "title": f"运行失败重试批次 {Path(batch_file).name}",
            "command": [py, "-u", "main_automation.py", "--update-list", batch_file],
            "env": {
                "USE_MITM_PROXY": "true",
                "MITM_PORT": str(MITM_PORT),
                "MITM_TIMEOUT": str(timeout),
                "CNIPA_LOGIN_WAIT_SECONDS": DEFAULT_LOGIN_WAIT_SECONDS,
            },
        }
    if action == "validate_results":
        return {"action": action, "title": "验证采集结果", "command": [py, "-u", "validate_results.py"]}
    if action == "analyze_recent":
        return {"action": action, "title": "分析采集状态", "command": [py, "-u", "analyze_collection_status.py"]}
    if action == "db_rebuild":
        # 从 detection_log.jsonl 重建 patents.db（DB 损坏或迁移场景）
        return {"action": action, "title": "从 JSONL 重建 DB", "command": [py, "-u", "sync.py", "rebuild"]}
    if action == "sync_status":
        return {"action": action, "title": "查看同步状态", "command": [py, "-u", "sync.py", "status"]}
    if action == "sync_pull_master":
        return {
            "action": action,
            "title": "从 master 拉取增量",
            "command": [py, "-u", "sync_pull_from_master.py"],
        }
    if action == "upgrade_code":
        return {"action": action, "title": "安全更新系统代码", "command": [py, "-u", "fetch_update.py"]}
    if action == "fetch_update":
        return {"action": action, "title": "无 git 更新代码", "command": [py, "-u", "fetch_update.py"]}
    if action == "check_update":
        return {"action": action, "title": "检查更新", "command": [py, "-u", "check_update.py"]}

    raise ValueError(f"未知操作: {action}")


def build_summary(job_manager: JobManager) -> dict[str, Any]:
    # 从 SQLite 获取所有聚合数据（替代多次 JSONL 全表扫描）
    db_summary = _patents_db.get_summary()

    # 策略分组（仍需读 focus_strategy.json + DB 记录）
    focus_strategy = safe_json_load(DATA_DIR / "focus_strategy.json", {})
    status_breakdown = focus_strategy.get("status_breakdown", {}) if isinstance(focus_strategy, dict) else {}

    # 策略分组计算（从 DB 读取最新一份记录用于计算 next_update）
    if status_breakdown:
        db_records_for_groups = _patents_db.get_all_records()
        update_groups = build_update_groups(db_records_for_groups, status_breakdown)
    else:
        update_groups = []

    search_app_nos = parse_app_no_list(safe_read_text(SEARCH_LIST_FILE))
    stored_app_nos = _patents_db.get_processed_app_nos()
    search_count = len(search_app_nos)
    search_collected = sum(1 for app_no in search_app_nos if app_no in stored_app_nos)
    dynamic_count = count_lines(DATA_DIR / "update_list_dynamic.txt")
    retry_count = count_lines(DATA_DIR / "retry_dynamic.txt")
    failed_retry_count = count_lines(RETRY_FAILED_FILE)
    config = safe_json_load(CONFIG_FILE, {})

    active_jobs = [j for j in job_manager.list_jobs() if j.get("status") in {"running", "stopping"}]

    warnings: list[str] = []
    if search_count == 0:
        warnings.append("申请号列表为空")
    if not CONFIG_FILE.exists():
        warnings.append("鼠标坐标配置不存在")
    elif config.get("input_x") == config.get("button_x") and config.get("input_y") == config.get("button_y"):
        warnings.append("输入框和查询按钮坐标相同，强制测试可能无法点击查询按钮")
    if dynamic_count == 0:
        warnings.append("动态更新清单为空")

    unique = db_summary["unique_count"]
    return {
        "now": iso_now(),
        "records": {
            "events": unique,
            "unique": unique,
            "success": db_summary["success"],
            "failed": db_summary["failed"],
            "pending": db_summary["pending"],
            "success_rate": db_summary["success_rate"],
            "bad_lines": 0,
        },
        "business": {
            "rejection": db_summary["rejection"],
            "fwxx_collected": db_summary["fwxx_collected"],
            "fwxx_pending": db_summary["fwxx_pending"],
            "tracked_total": sum(group["total"] for group in update_groups),
            "update_due": sum(group["due"] for group in update_groups),
        },
        "lists": {
            "search": search_count,
            "search_collected": search_collected,
            "dynamic": dynamic_count,
            "retry": retry_count,
            "failed_retry": failed_retry_count,
        },
        "proxy": {
            "host": MITM_HOST,
            "port": MITM_PORT,
            "reachable": port_open(MITM_HOST, MITM_PORT),
        },
        "config": config,
        "update_groups": update_groups,
        "status_counts": db_summary["status_counts"],
        "applicant_counts": db_summary["applicant_counts"],
        "recent": db_summary["recent"],
        "files": {key: file_info(path) for key, path in DOWNLOADS.items()},
        "jobs": active_jobs,
        "warnings": warnings,
        "daily_counts": db_summary["daily_counts"],
        "fwxx_pending_list": db_summary["fwxx_pending_list"],
        "pending_requests_count": len(_patents_db.list_requests(status='pending')),
        # 驳回企业列表，合并手动补录的真实专利数
        "rejection_companies": _merge_company_meta(db_summary["rejection_companies"]),
    }


def _merge_company_meta(rejection_companies: list[dict]) -> list[dict]:
    """将 DB 驳回企业列表与 company_meta.json 的手动补录数据合并。"""
    meta = safe_json_load(COMPANY_META_FILE, {})
    result = []
    for item in rejection_companies:
        name = item["name"]
        real_total = meta.get(name, {}).get("real_total") if isinstance(meta.get(name), dict) else None
        result.append({
            "name": name,
            "invention_count": item["invention_count"],
            "real_total": real_total,
        })
    return result


def build_update_groups(records: list[dict[str, Any]], status_breakdown: dict[str, Any]) -> list[dict[str, Any]]:
    groups: dict[int, dict[str, Any]] = {}
    now = utc_now()
    for status, info in status_breakdown.items():
        freq = int(info.get("frequency_days", 0) or 0)
        if not freq:
            continue
        group = groups.setdefault(
            freq,
            {
                "frequency_days": freq,
                "frequency_name": info.get("frequency_name") or f"{freq}天检查",
                "statuses": [],
                "total": 0,
                "due": 0,
                "waiting": 0,
                "earliest": None,
            },
        )
        group["statuses"].append(status)

    for item in records:
        status = item.get("anjianywzt")
        info = status_breakdown.get(status)
        if not info:
            continue
        freq = int(info.get("frequency_days", 0) or 0)
        if freq not in groups:
            continue
        group = groups[freq]
        group["total"] += 1
        last_update = parse_timestamp(item.get("timestamp"))
        next_update = last_update + timedelta(days=freq) if last_update else None
        needs_update = next_update is None or now >= next_update
        if needs_update:
            group["due"] += 1
        else:
            group["waiting"] += 1
        if next_update and (group["earliest"] is None or next_update.isoformat() < group["earliest"]):
            group["earliest"] = next_update.isoformat().replace("+00:00", "Z")

    return [groups[key] for key in sorted(groups)]


# ══════════════════════════════════════════════════════════════════════
#  前端资源
# ══════════════════════════════════════════════════════════════════════

HTML = r"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>CNIPA 采集控制台</title>
  <link rel="stylesheet" href="/app.css">
</head>
<body>
<div class="layout">

  <!-- ── 侧边栏 ── -->
  <nav class="sidebar">
    <div class="sidebar-logo">
      <span>CNIPA</span>
      <small>采集控制台</small>
    </div>
    <a class="nav-item" data-tab="overview">               <span>📊</span>概览</a>
    <a class="nav-item operator-only" data-tab="collection"><span>⚡</span>采集控制</a>
    <a class="nav-item" data-tab="strategy">               <span>📅</span>策略管理</a>
    <a class="nav-item" data-tab="fwxx">                   <span>📋</span>发文采集</a>
    <a class="nav-item operator-only" data-tab="public">   <span>🔍</span>公开查询</a>
    <a class="nav-item" data-tab="analytics">              <span>📈</span>数据分析</a>
    <a class="nav-item operator-only" data-tab="data">     <span>🗄</span>数据管理</a>
    <a class="nav-item operator-only" data-tab="logs">     <span>📟</span>任务日志</a>
    <a class="nav-item operator-only" data-tab="config">   <span>⚙</span>系统配置</a>
    <a class="nav-item viewer-only"   data-tab="requests"> <span>📥</span>提交需求</a>
  </nav>

  <!-- ── 主内容区 ── -->
  <div class="main">

    <!-- 顶部栏 -->
    <header class="topbar">
      <div class="top-left">
        <span id="clock">--</span>
        <span class="dot"></span>
        <span id="machineRolePill" class="pill muted">角色检测中</span>
        <span class="dot"></span>
        <span id="proxyPill" class="pill muted">代理检测中</span>
      </div>
      <div class="top-actions">
        <button class="btn secondary" data-action="export_excel">导出 Excel</button>
        <button class="btn secondary" data-action="export_json">导出 JSON</button>
        <button class="btn primary"   data-action="mitm_proxy">启动主代理</button>
      </div>
    </header>

    <section id="warnings" class="warnings hidden"></section>
    <section id="loginBanner" class="login-banner hidden">
      <span>🔐 浏览器正在等待您完成验证码并登录</span>
      <button class="btn primary" id="loginDoneBtn">✅ 我已完成验证码</button>
    </section>
    <section id="updateBanner" class="login-banner hidden">
      <span id="updateBannerText">🆕 发现新版本</span>
      <button class="btn primary" id="updateNowBtn">立即更新</button>
      <button class="btn secondary" id="updateDismissBtn">稍后</button>
    </section>

    <!-- ═══ Tab 1：概览 ═══ -->
    <div id="tab-overview" class="tab-panel">
      <section class="metrics">
        <article class="metric"><span>唯一申请号</span><strong id="mUnique">0</strong><em id="mEvents">0 条记录</em></article>
        <article class="metric"><span>成功率</span><strong id="mRate">0%</strong><em id="mSuccess">0 成功 / 0 失败</em></article>
        <article class="metric"><span>驳回目标</span><strong id="mRejection">0</strong><em id="mFwxx">0 待补发文</em></article>
        <article class="metric"><span>现在应检查</span><strong id="mDue">0</strong><em id="mTracked">0 跟踪中</em></article>
        <article class="metric"><span>动态清单</span><strong id="mDynamic">0</strong><em id="mSearch">0 输入申请号</em></article>
      </section>

      <section class="grid two">
        <article class="panel">
          <div class="panel-head"><h2>近 7 天采集量</h2><span class="hint">写入事件数</span></div>
          <svg id="trendSvg" viewBox="0 0 360 72" style="width:100%;height:72px;display:block;overflow:visible"></svg>
          <div id="trendLabels" class="trend-labels"></div>
        </article>
        <article class="panel">
          <div class="panel-head"><h2>系统状态</h2></div>
          <div class="info-grid">
            <div class="info-row"><span>MITM 代理</span><span id="healthProxy">—</span></div>
            <div class="info-row"><span>最后采集</span><span id="healthLastCollect">—</span></div>
            <div class="info-row"><span>日志大小</span><span id="healthLogSize">—</span></div>
            <div class="info-row"><span>Excel 文件</span><span id="healthExcel">—</span></div>
          </div>
        </article>
      </section>

      <article class="panel" style="margin-bottom:14px">
        <div class="panel-head"><h2>快捷操作</h2><span class="hint">常用</span></div>
        <div class="button-row">
          <button class="btn primary"   data-action="main_full">继续全量采集</button>
          <button class="btn secondary" data-action="collect_fwxx">补采发文</button>
          <button class="btn secondary" data-action="strategy_generate">生成策略清单</button>
          <button class="btn secondary" data-action="export_excel">导出 Excel</button>
        </div>
      </article>

      <article class="panel">
        <div class="panel-head"><h2>运行中任务</h2><span class="hint">实时</span></div>
        <div id="activeJobs"><span class="hint">暂无运行中的任务</span></div>
      </article>

      <article class="panel operator-only" style="margin-top:14px">
        <div class="panel-head"><h2>需求队列</h2><span class="hint" id="reqQueueHint">—</span></div>
        <div id="reqQueueList"><span class="hint">暂无待处理需求</span></div>
      </article>
    </div>

    <!-- ═══ Tab 2：采集控制 ═══ -->
    <div id="tab-collection" class="tab-panel">
      <article class="panel" style="margin-bottom:14px">
        <div class="panel-head"><h2>采集进度</h2><span class="hint" id="collectProgressHint">—</span></div>
        <div class="prog-bar-wrap"><div class="prog-bar" id="collectProgBar" style="width:0%"></div></div>
      </article>

      <section class="grid two">
        <article class="panel">
          <div class="panel-head"><h2>主流程采集</h2><span class="hint">main_automation</span></div>
          <div class="control-grid" style="margin-bottom:10px">
            <label class="field"><span>测试条数</span><input id="testCount" type="number" min="1" max="10000" value="5"></label>
            <button class="btn primary"   id="runTest">强制测试</button>
            <button class="btn secondary" id="collectFwxxTest">补采测试</button>
          </div>
          <div class="button-row" style="margin-bottom:14px">
            <button class="btn danger-soft" data-action="main_full">继续全量采集</button>
            <button class="btn secondary"   data-action="phase0_browser">Phase 0 浏览器</button>
            <button class="btn secondary"   data-action="import_cache">导入缓存</button>
          </div>
          <div class="panel-head"><h2>按清单更新</h2></div>
          <div class="control-grid">
            <label class="field">
              <span>更新清单</span>
              <select id="updateFile">
                <option value="data/update_list_dynamic.txt">动态清单</option>
                <option value="data/retry_dynamic.txt">动态重试清单</option>
                <option value="data/retry_failed.txt">失败重试清单</option>
              </select>
            </label>
            <label class="field"><span>限制条数</span><input id="updateLimit" type="number" min="1" max="10000" placeholder="不限制"></label>
            <button class="btn primary" id="runUpdate">按清单更新</button>
          </div>
        </article>

        <article class="panel">
          <div class="panel-head">
            <h2>申请号列表</h2>
            <button class="btn primary" id="saveSearchList">保存</button>
          </div>
          <textarea id="searchList" spellcheck="false"></textarea>
        </article>
      </section>
    </div>

    <!-- ═══ Tab 3：策略管理 ═══ -->
    <div id="tab-strategy" class="tab-panel">
      <article class="panel" style="margin-bottom:14px">
        <div class="panel-head"><h2>策略总览</h2></div>
        <div class="info-grid two-col">
          <div class="info-row"><span>跟踪申请数</span><strong id="stratTracked">—</strong></div>
          <div class="info-row"><span>现在应检查</span><strong id="stratDue">—</strong></div>
        </div>
      </article>

      <section class="grid two">
        <article class="panel">
          <div class="panel-head"><h2>策略分组</h2><span class="hint">按检查周期，点击"采集"立即触发</span></div>
          <div id="strategyGroups" class="group-list"></div>
        </article>

        <article class="panel operator-only">
          <div class="panel-head"><h2>辅助操作</h2></div>
          <div class="control-grid" style="margin-bottom:12px">
            <label class="field">
              <span>周期</span>
              <select id="strategyFrequency">
                <option value="">全部</option>
                <option value="7">7 天</option>
                <option value="14">14 天</option>
                <option value="30">30 天</option>
                <option value="45">45 天</option>
                <option value="60">60 天</option>
              </select>
            </label>
            <button class="btn primary"   id="generateStrategy">生成清单</button>
            <button class="btn secondary" id="statusStrategy">查看状态</button>
          </div>
          <div class="button-row" style="margin-bottom:12px">
            <button class="btn secondary" data-action="strategy_prepare">保存快照</button>
            <button class="btn secondary" data-action="strategy_diff">状态变化</button>
            <button class="btn secondary" data-action="strategy_report">详细报告</button>
            <button class="btn secondary" data-action="strategy_validate">校验策略</button>
            <button class="btn secondary" data-action="strategy_stats">策略统计</button>
          </div>
          <div class="check-line">
            <input id="singleAppNo" placeholder="输入申请号">
            <button class="btn primary" id="checkApp">单号判断</button>
          </div>
        </article>
      </section>
    </div>

    <!-- ═══ Tab 4：发文采集 ═══ -->
    <div id="tab-fwxx" class="tab-panel">
      <section class="grid two" style="margin-bottom:14px">
        <article class="panel ring-panel">
          <div class="ring-wrap">
            <svg class="ring-svg" viewBox="0 0 100 100">
              <circle cx="50" cy="50" r="40" class="ring-bg"/>
              <circle cx="50" cy="50" r="40" id="fwxxRingFg" class="ring-fg"/>
            </svg>
            <div class="ring-label">
              <strong id="fwxxPct">0%</strong>
              <span>发文完成率</span>
            </div>
          </div>
          <div class="ring-stats">
            <div class="info-row"><span>驳回案件</span><strong id="fwxxRejection">—</strong></div>
            <div class="info-row"><span>已采集发文</span><strong id="fwxxCollected">—</strong></div>
            <div class="info-row"><span>待补采</span><strong id="fwxxPending">—</strong></div>
          </div>
        </article>

        <article class="panel operator-only">
          <div class="panel-head"><h2>采集操作</h2><span class="hint">collect_fwxx</span></div>
          <div class="button-row" style="margin-bottom:14px">
            <button class="btn primary"   data-action="collect_fwxx">全量补采</button>
            <button class="btn secondary" id="fwxxTestBtn">测试 5 条</button>
          </div>
          <div class="check-line">
            <input id="fwxxAppNo" placeholder="单号采集：输入申请号">
            <button class="btn primary" id="fwxxSingleBtn">采集</button>
          </div>
        </article>
      </section>

      <article class="panel">
        <div class="panel-head"><h2>待补采列表</h2><span class="hint" id="fwxxPendingHint">最新 20 条</span></div>
        <div class="table-wrap">
          <table>
            <thead><tr><th>申请号</th><th>业务状态</th><th>最后采集</th></tr></thead>
            <tbody id="fwxxPendingRows"></tbody>
          </table>
        </div>
      </article>
    </div>

    <!-- ═══ Tab 5：公开查询 ═══ -->
    <div id="tab-public" class="tab-panel">
      <div class="steps">
        <article class="panel step-panel">
          <div class="step-num">1</div>
          <div style="flex:1">
            <h3 style="margin-bottom:10px">启动公开查询代理</h3>
            <button class="btn primary" data-action="public_mitm_proxy">公开代理</button>
          </div>
        </article>
        <article class="panel step-panel">
          <div class="step-num">2</div>
          <div style="flex:1">
            <h3 style="margin-bottom:10px">打开浏览器并翻页</h3>
            <div class="button-row" style="margin-bottom:12px">
              <button class="btn secondary" data-action="public_browser">公开浏览器</button>
            </div>
            <div class="control-grid">
              <label class="field"><span>翻页间隔（秒）</span><input id="pageDelay" type="number" min="0.2" max="30" step="0.1" value="1.5"></label>
              <label class="field"><span>最大页数</span><input id="maxPages" type="number" min="1" max="10000" value="50"></label>
              <button class="btn primary" id="autoPaginate">自动翻页</button>
            </div>
            <div style="margin-top:10px;padding:10px;background:var(--bg2);border-radius:6px;font-size:13px">
              启动后在浏览器中设置查询条件并点击查询，出现结果后点击下方按钮：
              <div style="margin-top:8px">
                <button class="btn secondary" id="signalQueryReady">✅ 我已完成查询设置</button>
              </div>
            </div>
          </div>
        </article>
        <article class="panel step-panel">
          <div class="step-num">3</div>
          <div style="flex:1">
            <h3 style="margin-bottom:10px">导出 &amp; 入库</h3>
            <div class="button-row" style="margin-bottom:12px">
              <button class="btn secondary" data-action="public_export">导出公开结果</button>
              <button class="btn primary"   data-action="import_public_search">导入系统库</button>
            </div>
            <div style="font-size:12px;color:var(--text2);margin-bottom:10px">
              「导入系统库」将采集数据写入 patents.db，可在概览和数据分析中查看
            </div>
            <div class="downloads">
              <a href="/download/excel">Excel ↓</a>
              <a href="/download/jsonl">JSONL ↓</a>
              <a href="/download/json">JSON ↓</a>
              <a href="/download/dynamic">动态清单 ↓</a>
            </div>
          </div>
        </article>
      </div>
    </div>

    <!-- ═══ Tab 6：数据分析 ═══ -->
    <div id="tab-analytics" class="tab-panel">
      <section class="grid two" style="margin-bottom:14px">
        <article class="panel">
          <div class="panel-head"><h2>业务状态分布</h2><span class="hint">TOP 12</span></div>
          <div id="statusCounts" class="bar-list"></div>
        </article>
        <article class="panel">
          <div class="panel-head"><h2>申请人分布</h2><span class="hint">TOP 8</span></div>
          <div id="applicantCounts" class="bar-list"></div>
        </article>
      </section>
      <article class="panel" style="margin-bottom:14px">
        <div class="panel-head">
          <h2>驳回企业列表</h2>
          <span class="hint">仅「驳回等复审请求」发明专利</span>
          <div style="margin-left:auto;display:flex;gap:6px">
            <button class="btn secondary" style="font-size:12px;padding:3px 10px" onclick="downloadCompanyTemplate()">下载模板</button>
            <label class="btn secondary" style="font-size:12px;padding:3px 10px;cursor:pointer;margin:0">
              上传补录
              <input type="file" accept=".xlsx,.xls" style="display:none" onchange="uploadCompanyMeta(this)">
            </label>
          </div>
        </div>
        <div style="margin-bottom:8px">
          <input id="rejCompanySearch" type="text" placeholder="过滤企业名…"
                 style="width:100%;box-sizing:border-box;padding:5px 8px;border:1px solid var(--line);border-radius:4px;font-size:13px">
        </div>
        <div class="table-wrap">
          <table id="rejCompanyTable">
            <thead>
              <tr>
                <th>企业名</th>
                <th style="width:90px;text-align:right">库内发明数</th>
                <th style="width:130px;text-align:right">实际专利总数</th>
                <th style="width:60px"></th>
              </tr>
            </thead>
            <tbody id="rejCompanyRows"><tr><td colspan="4" class="hint" style="padding:8px">加载中…</td></tr></tbody>
          </table>
        </div>
      </article>

      <article class="panel" style="margin-bottom:14px">
        <div class="panel-head"><h2>验证与分析</h2></div>
        <div class="button-row">
          <button class="btn primary"   data-action="validate_results">运行验证</button>
          <button class="btn secondary" data-action="analyze_recent">采集状态分析</button>
        </div>
      </article>
      <article class="panel" style="margin-bottom:14px">
        <div class="panel-head"><h2>筛选导出</h2><span class="hint">条件为「或」关系</span></div>
        <div class="control-grid" style="margin-bottom:10px">
          <label class="field" style="grid-column:1/-1">
            <span>企业 / 申请人（可多选，输入关键词过滤）</span>
            <input id="exportApplicantSearch" type="text" placeholder="输入关键词过滤下方列表，如「三花」">
          </label>
        </div>
        <div id="exportApplicantList"
             style="max-height:160px;overflow-y:auto;border:1px solid var(--line);border-radius:6px;padding:8px;margin-bottom:12px;font-size:13px">
          <span class="hint">加载申请人列表中...</span>
        </div>
        <div class="control-grid">
          <label class="field"><span>采集时间 起</span><input id="exportTsFrom" type="date"></label>
          <label class="field"><span>采集时间 止</span><input id="exportTsTo" type="date"></label>
          <label class="field"><span>驳回发文日期 起</span><input id="exportRejFrom" type="date"></label>
          <label class="field"><span>驳回发文日期 止</span><input id="exportRejTo" type="date"></label>
        </div>
        <div class="hint" style="margin-top:8px">
          满足任一条件即导出（如选企业 + 驳回日期范围 = 该企业的全部 ∪ 该日期范围内被驳回的全部）。
          全部留空 = 导出全量。
        </div>
        <div class="button-row" style="margin-top:10px;align-items:center">
          <button class="btn secondary" id="exportPreviewBtn">预估数量</button>
          <button class="btn primary" id="exportFilteredBtn">📥 导出 Excel</button>
          <span id="exportFilterHint" class="hint" style="margin-left:8px"></span>
        </div>
      </article>
      <article class="panel">
        <div class="panel-head"><h2>最近记录</h2><span class="hint">最新写入 16 条</span></div>
        <div class="table-wrap">
          <table>
            <thead>
              <tr><th>申请号</th><th>状态码</th><th>业务状态</th><th>专利名称</th><th>申请人</th><th>耗时</th><th>时间</th></tr>
            </thead>
            <tbody id="recentRows"></tbody>
          </table>
        </div>
      </article>
    </div>

    <!-- ═══ Tab 7：数据管理 ═══ -->
    <div id="tab-data" class="tab-panel">
      <section class="grid two" style="margin-bottom:14px">
        <article class="panel">
          <div class="panel-head"><h2>重试管理</h2><span class="hint">retry_failed</span></div>
          <div class="info-grid" style="margin-bottom:14px">
            <div class="info-row"><span>历史失败</span><span id="retryHistoryFailed">—</span></div>
            <div class="info-row"><span>重试清单</span><span id="retryCount">—</span></div>
          </div>
          <div class="control-grid" style="margin-bottom:10px">
            <label class="field">
              <span>批次数量</span>
              <input id="retryBatchSize" type="number" min="1" max="10000" value="200">
            </label>
            <label class="field">
              <span>超时秒数</span>
              <input id="retryTimeout" type="number" min="1" max="120" value="12">
            </label>
          </div>
          <div class="button-row">
            <button class="btn secondary" data-action="retry_failed">生成失败清单</button>
            <button class="btn secondary" id="retryRecollectBtn">立即重新采集</button>
            <button class="btn secondary" id="retryBatchBtn">生成批次</button>
            <button class="btn primary" id="runRetryBatchBtn">运行批次</button>
          </div>
          <div class="hint" style="margin-top:6px">历史失败是数据库累计值；立即重采需 MITM 主代理已启动；批次默认写入 data/retry_batch_001.txt</div>
        </article>
        <article class="panel">
          <div class="panel-head"><h2>数据库维护</h2></div>
          <div class="info-grid" style="margin-bottom:14px">
            <div class="info-row"><span>JSONL 日志</span><span id="jsonlSize">—</span></div>
          </div>
          <div class="button-row">
            <button class="btn secondary" data-action="db_rebuild">从 JSONL 重建 DB</button>
          </div>
          <div class="hint" style="margin-top:6px">patents.db 损坏或迁移时，从 detection_log.jsonl 重建</div>
        </article>
      </section>
      <article class="panel" style="margin-bottom:14px">
        <div class="panel-head"><h2>多机同步</h2><span class="hint">master → replica</span></div>
        <div class="button-row">
          <button class="btn secondary" data-action="sync_status">查看同步状态</button>
          <button class="btn primary" data-action="sync_pull_master">从 master 拉取增量</button>
        </div>
        <div class="hint" style="margin-top:8px">副本机只从部署机 Dashboard 拉取增量；成功后自动创建数据提交，再由人工执行 git push</div>
      </article>
      <article class="panel operator-only" style="margin-bottom:14px">
        <div class="panel-head"><h2>代理机构导入</h2><span class="hint">CSV / Excel → patents.db</span></div>
        <div class="control-grid">
          <label class="field">
            <span>选择名单文件</span>
            <input id="agencyFileInput" type="file" accept=".csv,.xlsx,.xls">
          </label>
          <button class="btn primary" id="importAgencyBtn">导入代理机构</button>
        </div>
        <div class="hint" style="margin-top:6px">
          文件需包含「申请号」和「代理机构」两列（支持中英文列名）；
          发文专利的代理机构由 MITM 自动采集，无需手动上传。
        </div>
        <div id="agencyImportHint" class="hint" style="margin-top:6px"></div>
      </article>
      <article class="panel operator-only">
        <div class="panel-head"><h2>增量数据互通</h2><span class="hint">跨机最小化传输</span></div>
        <section class="grid two">
          <div>
            <div class="panel-head" style="margin-bottom:8px"><h3 style="font-size:13px">导出（发给其他机器）</h3></div>
            <div class="control-grid">
              <label class="field">
                <span>起始时间</span>
                <input id="deltaFrom" type="datetime-local">
              </label>
              <button class="btn primary" id="exportDeltaBtn">下载增量 JSONL</button>
            </div>
            <div id="deltaExportHint" class="hint" style="margin-top:6px"></div>
          </div>
          <div>
            <div class="panel-head" style="margin-bottom:8px"><h3 style="font-size:13px">导入（接收其他机器的数据）</h3></div>
            <div class="control-grid">
              <label class="field">
                <span>选择 JSONL 文件</span>
                <input id="deltaFileInput" type="file" accept=".jsonl,.json">
              </label>
              <button class="btn primary" id="importDeltaBtn">导入合并</button>
            </div>
            <div id="deltaImportHint" class="hint" style="margin-top:6px"></div>
          </div>
        </section>
      </article>
    </div>

    <!-- ═══ Tab 8：任务日志 ═══ -->
    <div id="tab-logs" class="tab-panel">
      <section class="grid wide-left">
        <article class="panel">
          <div class="panel-head">
            <h2>日志输出</h2>
            <div style="display:flex;gap:8px;align-items:center">
              <button class="btn primary hidden" id="resumeLoginBtn">✅ 我已完成验证码</button>
              <button class="btn secondary" id="stopJob">停止</button>
            </div>
          </div>
          <pre id="jobLog" class="terminal">等待任务启动...</pre>
        </article>
        <article class="panel">
          <div class="panel-head"><h2>任务列表</h2></div>
          <div id="jobList" class="job-list"></div>
        </article>
      </section>
    </div>

    <!-- ═══ Tab 9：系统配置 ═══ -->
    <div id="tab-config" class="tab-panel">
      <section class="grid two" style="margin-bottom:14px">
        <article class="panel">
          <div class="panel-head">
            <h2>鼠标坐标配置</h2>
            <div class="button-row">
              <button class="btn secondary" id="resetConfig">重录坐标</button>
              <button class="btn primary"   id="saveConfig">保存配置</button>
            </div>
          </div>
          <textarea id="configText" class="codebox" spellcheck="false"></textarea>
        </article>
        <article class="panel">
          <div class="panel-head"><h2>系统信息</h2><span class="hint">只读</span></div>
          <div class="info-grid">
            <div class="info-row"><span>MITM 代理</span><span id="sysProxy">—</span></div>
            <div class="info-row"><span>JSONL 大小</span><span id="sysJsonlSize">—</span></div>
            <div class="info-row"><span>动态清单</span><span id="sysDynamic">—</span></div>
            <div class="info-row"><span>重试清单</span><span id="sysRetry">—</span></div>
          </div>
          <div class="field" style="margin-top:14px">
            <span>远程写操作 Token</span>
            <input id="apiTokenInput" type="password" placeholder="本机自动载入；远程操作时粘贴">
          </div>
          <button class="btn secondary" id="saveApiToken" style="margin-top:8px">保存到此浏览器</button>
        </article>
      </section>
      <article class="panel operator-only" style="margin-bottom:14px">
        <div class="panel-head"><h2>代码更新</h2><span class="hint">check_update / upgrade / fetch_update</span></div>
        <div class="button-row">
          <button class="btn secondary" id="checkUpdateBtn">🔎 检查更新</button>
          <button class="btn primary" data-action="upgrade_code">🔄 更新系统代码（git）</button>
          <button class="btn secondary" data-action="fetch_update">📥 无 git 更新（HTTP）</button>
        </div>
        <div class="hint" style="margin-top:6px">发现新版本时顶部会自动提示；有网络时从 GitHub 拉取最新代码，无 git 环境可用 HTTP 方式</div>
      </article>
      <article class="panel operator-only">
        <div class="panel-head">
          <h2>登录凭证</h2>
          <button class="btn primary" id="saveCreds">保存</button>
        </div>
        <div class="control-grid">
          <label class="field">
            <span>代理机构代码</span>
            <input id="credsUser" type="text" placeholder="CNIPA_USERNAME" autocomplete="username">
          </label>
          <label class="field">
            <span>密码</span>
            <input id="credsPass" type="password" placeholder="CNIPA_PASSWORD（留空则不修改）" autocomplete="current-password">
          </label>
        </div>
        <div id="credsStatus" class="hint" style="margin-top:8px"></div>
      </article>
    </div>

    <!-- ═══ Tab 10：提交需求 ═══ -->
    <div id="tab-requests" class="tab-panel">
      <article class="panel" style="margin-bottom:14px">
        <div class="panel-head"><h2>提交采集需求</h2><span class="hint">每行一个申请号</span></div>
        <div class="control-grid" style="margin-bottom:12px">
          <label class="field" style="grid-column:1/-1">
            <span>申请号列表</span>
            <textarea id="reqAppNos" class="codebox" rows="8" placeholder="每行填写一个申请号，例如：&#10;2023117765870&#10;2022108928573"></textarea>
          </label>
          <label class="field">
            <span>备注（可选）</span>
            <input id="reqNote" type="text" placeholder="说明采集原因或来源">
          </label>
          <button class="btn primary" id="submitReqBtn">提交需求</button>
        </div>
        <div id="reqSubmitResult" class="hint" style="margin-top:8px"></div>
      </article>
      <article class="panel viewer-only">
        <div class="panel-head"><h2>我的提交记录</h2></div>
        <div id="myReqList"><span class="hint">暂无记录</span></div>
      </article>
    </div>

  </div><!-- /main -->
</div><!-- /layout -->

<div id="toast" class="toast hidden"></div>
<script src="/app.js"></script>
</body>
</html>
"""


CSS = r""":root {
  color-scheme: light;
  --bg: #eef1ef;
  --panel: #ffffff;
  --panel-soft: #f7f8f5;
  --ink: #202623;
  --muted: #66736d;
  --line: #d9ded8;
  --accent: #147a63;
  --accent-dark: #0e5d4d;
  --amber: #ad6b00;
  --red: #b13b3b;
  --shadow: 0 4px 20px rgba(31,42,36,.07);
  --sidebar-w: 174px;
}

* { box-sizing: border-box; }

body {
  margin: 0;
  min-height: 100vh;
  font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont,
               "Segoe UI", "Microsoft YaHei", sans-serif;
  color: var(--ink);
  background: var(--bg);
}

button, input, select, textarea { font: inherit; }

/* ── Layout ── */
.layout { display: flex; min-height: 100vh; }

.sidebar {
  width: var(--sidebar-w);
  flex-shrink: 0;
  position: fixed;
  top: 0; left: 0; bottom: 0;
  background: var(--panel);
  border-right: 1px solid var(--line);
  display: flex;
  flex-direction: column;
  overflow-y: auto;
  z-index: 100;
}

.sidebar-logo {
  padding: 20px 16px 14px;
  border-bottom: 1px solid var(--line);
  margin-bottom: 8px;
}
.sidebar-logo span { display: block; font-weight: 800; font-size: 15px; color: var(--accent); }
.sidebar-logo small { color: var(--muted); font-size: 11px; }

.nav-item {
  display: flex;
  align-items: center;
  gap: 8px;
  padding: 9px 14px;
  font-size: 13px;
  color: var(--muted);
  text-decoration: none;
  cursor: pointer;
  border-left: 3px solid transparent;
  transition: background .1s, color .1s;
  user-select: none;
  white-space: nowrap;
}
.nav-item span { font-size: 14px; line-height: 1; }
.nav-item:hover { background: var(--bg); color: var(--ink); }
.nav-item.active {
  color: var(--accent-dark);
  background: #e9f5ef;
  border-left-color: var(--accent);
  font-weight: 600;
}

.main {
  margin-left: var(--sidebar-w);
  flex: 1;
  min-width: 0;
  padding: 0 24px 48px;
}

/* ── Top Bar ── */
.topbar {
  position: sticky;
  top: 0;
  z-index: 50;
  background: rgba(238,241,239,.95);
  backdrop-filter: blur(8px);
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 16px;
  padding: 11px 0;
  margin-bottom: 16px;
  border-bottom: 1px solid var(--line);
}

.top-left {
  display: flex;
  align-items: center;
  gap: 10px;
  font-size: 13px;
  color: var(--muted);
}

/* ── Tab Panels ── */
.tab-panel { display: none; }
.tab-panel.active { display: block; }
/* 角色控制：操作员默认显示，查看者模式下隐藏 */
.operator-only { display: block; }
.viewer-only   { display: none; }
body.viewer-mode .operator-only { display: none !important; }
body.viewer-mode .viewer-only   { display: block; }

/* ── Type ── */
h1, h2, h3, p { margin: 0; }
h2 { font-size: 16px; }
h3 { font-size: 13px; }
.hint { color: var(--muted); font-size: 12px; }

.dot { width: 4px; height: 4px; border-radius: 999px; background: #9aa59f; }

/* ── Pill ── */
.pill {
  display: inline-flex;
  align-items: center;
  min-height: 22px;
  padding: 0 8px;
  border-radius: 999px;
  border: 1px solid var(--line);
  background: #fff;
  color: var(--muted);
  font-size: 12px;
}
.pill.ok   { color: var(--accent-dark); background: #e9f5ef; border-color: #b9dcca; }
.pill.warn { color: var(--amber);       background: #fff7e8; border-color: #efd19c; }
.pill.master { color: #fff; background: var(--red); border-color: var(--red); font-weight: 700; }

/* ── Warnings ── */
.warnings {
  margin-bottom: 14px;
  border: 1px solid #efd19c;
  background: #fff8ec;
  color: #755014;
  border-radius: 8px;
  padding: 10px 14px;
  font-size: 13px;
}
.hidden { display: none !important; }

.login-banner {
  margin-bottom: 14px;
  border: 2px solid #147a63;
  background: #e6f5f0;
  color: #0e5d4d;
  border-radius: 8px;
  padding: 12px 16px;
  font-size: 14px;
  font-weight: 600;
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 16px;
  animation: bannerPulse 2s infinite;
}

@keyframes bannerPulse {
  0%, 100% { border-color: #147a63; }
  50% { border-color: #1fa882; box-shadow: 0 0 0 3px rgba(20,122,99,.15); }
}

/* ── Grid / Panel ── */
.grid { display: grid; gap: 14px; margin-bottom: 14px; }
.grid.two { grid-template-columns: 1fr 1fr; }
.grid.wide-left { grid-template-columns: minmax(0, 1.35fr) minmax(300px, .65fr); }

.panel {
  background: var(--panel);
  border: 1px solid var(--line);
  border-radius: 8px;
  padding: 16px;
  box-shadow: var(--shadow);
}

.panel-head {
  display: flex;
  justify-content: space-between;
  align-items: center;
  gap: 12px;
  margin-bottom: 14px;
}

/* ── Buttons ── */
.top-actions, .button-row, .downloads {
  display: flex;
  flex-wrap: wrap;
  gap: 8px;
  align-items: center;
}

.btn {
  border: 1px solid transparent;
  border-radius: 7px;
  min-height: 34px;
  padding: 0 13px;
  cursor: pointer;
  color: var(--ink);
  background: #edf1ec;
  font-size: 13px;
  white-space: nowrap;
  transition: transform .1s, background .1s;
}
.btn:active { transform: translateY(0) !important; }
.btn:hover  { transform: translateY(-1px); }
.btn.primary      { background: var(--accent); color: #fff; }
.btn.primary:hover { background: var(--accent-dark); }
.btn.secondary    { background: #f4f5f2; border-color: var(--line); }
.btn.danger-soft  { background: #fff2ef; border-color: #f0c8bd; color: #913729; }

/* ── Form ── */
input, select, textarea {
  width: 100%;
  border: 1px solid var(--line);
  border-radius: 7px;
  background: #fff;
  color: var(--ink);
  outline: none;
}
input, select { height: 34px; padding: 0 10px; }
textarea { min-height: 260px; resize: vertical; padding: 10px; line-height: 1.5; font-size: 13px; }
input:focus, select:focus, textarea:focus {
  border-color: rgba(20,122,99,.68);
  box-shadow: 0 0 0 3px rgba(20,122,99,.12);
}

.control-grid {
  display: grid;
  grid-template-columns: minmax(0,1fr) minmax(0,1fr) auto;
  gap: 10px;
  align-items: end;
  margin-bottom: 10px;
}

.field { display: grid; gap: 5px; }
.field span { color: var(--muted); font-size: 12px; }

.check-line {
  display: grid;
  grid-template-columns: minmax(0,1fr) auto;
  gap: 8px;
  margin-top: 10px;
}

.codebox {
  min-height: 200px;
  font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, "Liberation Mono", monospace;
  font-size: 12px;
}

/* ── Info Grid ── */
.info-grid { display: grid; gap: 0; }
.info-grid.two-col { grid-template-columns: 1fr 1fr; gap: 0 20px; }
.info-row {
  display: flex;
  justify-content: space-between;
  align-items: center;
  gap: 12px;
  padding: 7px 0;
  border-bottom: 1px solid #edf0ed;
  font-size: 13px;
}
.info-row:last-child { border-bottom: 0; }
.info-row > span:first-child { color: var(--muted); flex-shrink: 0; }

/* ── Metrics ── */
.metrics {
  display: grid;
  grid-template-columns: repeat(5, minmax(0,1fr));
  gap: 12px;
  margin-bottom: 14px;
}
.metric {
  background: var(--panel);
  border: 1px solid var(--line);
  border-radius: 8px;
  padding: 14px;
  box-shadow: var(--shadow);
}
.metric span   { display: block; color: var(--muted); font-size: 12px; margin-bottom: 7px; }
.metric strong { display: block; font-size: 28px; line-height: 1.1; }
.metric em     { display: block; color: var(--muted); font-style: normal; font-size: 11px; margin-top: 6px; }

/* ── Trend Chart ── */
.trend-labels {
  display: flex;
  justify-content: space-between;
  font-size: 11px;
  color: var(--muted);
  margin-top: 4px;
  padding: 0 6px;
}

/* ── Active Jobs (overview) ── */
#activeJobs { display: grid; gap: 8px; }
.active-job-item {
  display: flex;
  align-items: center;
  gap: 10px;
  padding: 8px 12px;
  border: 1px solid var(--line);
  border-radius: 7px;
  background: var(--panel-soft);
  font-size: 13px;
}
.running-dot {
  width: 8px; height: 8px;
  border-radius: 999px;
  background: var(--accent);
  flex-shrink: 0;
  animation: pulse 1.5s infinite;
}
@keyframes pulse { 0%,100%{opacity:1} 50%{opacity:.35} }

/* ── Progress Bar ── */
.prog-bar-wrap { height: 10px; background: #edf1ec; border-radius: 999px; overflow: hidden; }
.prog-bar { height: 100%; background: linear-gradient(90deg, var(--accent), #1fa882); border-radius: 999px; transition: width .6s ease; }

/* ── Strategy Groups ── */
.group-list { display: grid; gap: 10px; }
.group-item {
  display: grid;
  grid-template-columns: 72px minmax(0,1fr) 72px auto;
  gap: 10px;
  align-items: center;
  padding: 10px 0;
  border-top: 1px solid var(--line);
}
.group-item:first-child { border-top: 0; padding-top: 0; }
.bar { height: 7px; background: #edf1ec; border-radius: 999px; overflow: hidden; margin-top: 4px; }
.bar span { display: block; height: 100%; background: linear-gradient(90deg, var(--accent), #d69a38); }

/* ── Ring Chart (FWXX) ── */
.ring-panel { display: flex; align-items: center; gap: 20px; }
.ring-wrap { display: flex; align-items: center; gap: 14px; flex-shrink: 0; }
.ring-svg { width: 96px; height: 96px; transform: rotate(-90deg); }
.ring-bg { fill: none; stroke: #edf1ec; stroke-width: 10; }
.ring-fg {
  fill: none;
  stroke: var(--accent);
  stroke-width: 10;
  stroke-linecap: round;
  stroke-dasharray: 251.3;
  stroke-dashoffset: 251.3;
  transition: stroke-dashoffset .6s ease;
}
.ring-label { text-align: center; }
.ring-label strong { display: block; font-size: 20px; color: var(--accent-dark); }
.ring-label span { font-size: 11px; color: var(--muted); }
.ring-stats { flex: 1; min-width: 0; }

/* ── Public Search Steps ── */
.steps { display: grid; gap: 14px; }
.step-panel { display: flex; gap: 16px; align-items: flex-start; }
.step-num {
  width: 30px; height: 30px;
  border-radius: 999px;
  background: var(--accent);
  color: #fff;
  font-weight: 700; font-size: 14px;
  display: flex; align-items: center; justify-content: center;
  flex-shrink: 0;
  margin-top: 2px;
}

/* ── Bar List (Analytics) ── */
.bar-list { display: grid; gap: 8px; }
.bar-row {
  display: grid;
  grid-template-columns: minmax(100px,1.2fr) 100px 36px;
  align-items: center;
  gap: 8px;
  font-size: 12px;
}
.bar-row .name { overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.bar-thin { height: 5px; background: #edf1ec; border-radius: 999px; overflow: hidden; }
.bar-thin span { display: block; height: 100%; background: var(--accent); opacity: .75; border-radius: 999px; }
.bar-row .cnt { color: var(--muted); text-align: right; }

/* ── Table ── */
.table-wrap { overflow: auto; }
table { width: 100%; border-collapse: collapse; min-width: 560px; }
th, td { text-align: left; border-bottom: 1px solid var(--line); padding: 8px 8px; font-size: 13px; vertical-align: top; }
th { color: var(--muted); font-weight: 600; background: var(--panel-soft); }
td .clip { max-width: 260px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; display: block; }

/* ── Job List (Logs tab) ── */
.job-list { display: grid; gap: 6px; overflow-y: auto; max-height: 510px; }
.job-item {
  display: flex;
  align-items: center;
  gap: 10px;
  padding: 9px 11px;
  border: 1px solid var(--line);
  border-radius: 7px;
  cursor: pointer;
  background: var(--panel-soft);
  font-size: 13px;
  transition: background .1s;
}
.job-item:hover { background: #edf6f0; }
.job-item.selected { background: #e2f0e8; border-color: #b9dcca; }
.jdot { width: 8px; height: 8px; border-radius: 999px; background: var(--muted); flex-shrink: 0; }
.jdot.ok      { background: #6aae84; }
.jdot.running { background: var(--accent); animation: pulse 1.5s infinite; }
.jdot.err     { background: var(--red); }
.job-meta { min-width: 0; flex: 1; }
.job-meta strong { display: block; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; margin-bottom: 2px; }

/* ── Terminal ── */
.terminal {
  min-height: 490px;
  max-height: 560px;
  overflow: auto;
  margin: 0;
  padding: 13px;
  border-radius: 8px;
  background: #1f2421;
  color: #e8eee9;
  font: 13px/1.45 ui-monospace, SFMono-Regular, Menlo, Consolas, "Liberation Mono", monospace;
  white-space: pre-wrap;
}
.terminal .lo { color: #7ec89a; }
.terminal .le { color: #f07878; }
.terminal .lw { color: #f0c060; }

/* ── Downloads ── */
.downloads a {
  display: inline-flex;
  align-items: center;
  height: 30px;
  padding: 0 10px;
  border-radius: 7px;
  color: var(--accent-dark);
  background: #edf6f0;
  text-decoration: none;
  font-size: 12px;
  margin-right: 6px;
}

/* ── Toast ── */
.toast {
  position: fixed;
  right: 20px; bottom: 20px;
  max-width: min(420px, calc(100vw - 40px));
  padding: 11px 14px;
  border-radius: 8px;
  background: #202623;
  color: #fff;
  box-shadow: 0 8px 30px rgba(0,0,0,.25);
  font-size: 14px;
  z-index: 9999;
}

/* ── Responsive ── */
@media (max-width: 1100px) {
  .metrics { grid-template-columns: repeat(3, 1fr); }
  .grid.two, .grid.wide-left { grid-template-columns: 1fr; }
  .ring-panel { flex-direction: column; align-items: flex-start; }
  .info-grid.two-col { grid-template-columns: 1fr; }
}

@media (max-width: 720px) {
  .metrics { grid-template-columns: 1fr 1fr; }
  .control-grid, .check-line { grid-template-columns: 1fr; }
  .main { padding: 0 14px 40px; }
}
"""


JS = r"""const state = {
  currentTab: 'overview',
  selectedJobId: null,
  searchLoaded: false,
  configLoaded: false,
  roleDetermined: false,
  apiToken: localStorage.getItem('cnipaApiToken') || '',
};

const $ = (s) => document.querySelector(s);
const $$ = (s) => Array.from(document.querySelectorAll(s));

// ── Utilities ─────────────────────────────────────────────────────────
function fmtNumber(v) { return Number(v || 0).toLocaleString('zh-CN'); }

function fmtBytes(b) {
  if (!b) return '—';
  if (b < 1024) return b + ' B';
  if (b < 1048576) return (b / 1024).toFixed(1) + ' KB';
  return (b / 1048576).toFixed(1) + ' MB';
}

function shortTime(v) {
  if (!v) return '—';
  const d = new Date(v);
  return isNaN(d.getTime()) ? v : d.toLocaleString('zh-CN', { hour12: false });
}

function relTime(v) {
  if (!v) return '—';
  const diff = Date.now() - new Date(v).getTime();
  if (diff < 60000)    return '刚刚';
  if (diff < 3600000)  return Math.floor(diff / 60000) + ' 分钟前';
  if (diff < 86400000) return Math.floor(diff / 3600000) + ' 小时前';
  return Math.floor(diff / 86400000) + ' 天前';
}

function escHtml(v) {
  return String(v)
    .replaceAll('&', '&amp;').replaceAll('<', '&lt;')
    .replaceAll('>', '&gt;').replaceAll('"', '&quot;').replaceAll("'", '&#039;');
}

function colorLine(line) {
  const e = escHtml(line);
  if (/[✓✅]|成功|[Ss]uccess|finished|完成|采集成功/.test(line)) return '<span class="lo">' + e + '</span>';
  if (/[✗❌]|错误|[Ee]rror|[Ff]ail|失败|异常|Traceback|Exception|CRITICAL/.test(line)) return '<span class="le">' + e + '</span>';
  if (/[⚠]|警告|[Ww]arn|重试|retry|跳过|SKIP/.test(line)) return '<span class="lw">' + e + '</span>';
  return e;
}

function showToast(msg) {
  const t = $('#toast');
  t.textContent = msg;
  t.classList.remove('hidden');
  clearTimeout(showToast._t);
  showToast._t = setTimeout(() => t.classList.add('hidden'), 2800);
}

async function api(path, opts = {}) {
  const headers = { 'Content-Type': 'application/json', ...(opts.headers || {}) };
  if (state.apiToken) headers['X-CNIPA-Token'] = state.apiToken;
  const res = await fetch(path, { ...opts, headers });
  const text = await res.text();
  let payload = {};
  try { payload = text ? JSON.parse(text) : {}; } catch { payload = { raw: text }; }
  if (!res.ok) throw new Error(payload.error || res.statusText);
  return payload;
}

function writeHeaders(extra = {}) {
  const headers = { ...extra };
  if (state.apiToken) headers['X-CNIPA-Token'] = state.apiToken;
  return headers;
}

async function loadOperatorToken() {
  try {
    const res = await fetch('/api/operator-token');
    if (res.ok) {
      const payload = await res.json();
      state.apiToken = payload.token;
      localStorage.setItem('cnipaApiToken', state.apiToken);
    }
  } catch (_) {}
  const input = $('#apiTokenInput');
  if (input) input.value = state.apiToken;
}

// ── Tab Routing ───────────────────────────────────────────────────────
function switchTab(tab) {
  $$('.tab-panel').forEach(p => p.classList.remove('active'));
  $$('.nav-item').forEach(n => n.classList.remove('active'));
  const panel = document.getElementById('tab-' + tab);
  if (panel) panel.classList.add('active');
  const nav = document.querySelector('.nav-item[data-tab="' + tab + '"]');
  if (nav) nav.classList.add('active');
  location.hash = tab;
  state.currentTab = tab;
}

function initTabRouting() {
  $$('.nav-item').forEach(item => item.addEventListener('click', () => switchTab(item.dataset.tab)));
  const hash = location.hash.replace('#', '') || 'overview';
  switchTab(hash);
}

// ── Job Control ───────────────────────────────────────────────────────
async function startJob(action, params = {}) {
  try {
    const data = await api('/api/jobs', { method: 'POST', body: JSON.stringify({ action, params }) });
    state.selectedJobId = data.job.id;
    showToast('已启动：' + data.job.title);
    switchTab('logs');
    await refreshJobs();
  } catch (e) { showToast('启动失败：' + e.message); }
}

async function refreshJobs() {
  const data = await api('/api/jobs');
  const jobs = data.jobs || [];
  if (!state.selectedJobId && jobs.length) state.selectedJobId = jobs[0].id;
  renderJobList(jobs);
  await refreshJobLog();
  // 全局登录横幅：任意运行中任务等待验证码时显示
  const needLogin = jobs.some(j => j.waiting_for_login);
  const banner = $('#loginBanner');
  if (banner) banner.classList.toggle('hidden', !needLogin);
}

async function refreshJobLog() {
  if (!state.selectedJobId) { $('#jobLog').textContent = '等待任务启动...'; return; }
  try {
    const data = await api('/api/jobs/' + state.selectedJobId);
    const logs = data.job.logs || [];
    const term = $('#jobLog');
    term.innerHTML = logs.map(colorLine).join('\n');
    term.scrollTop = term.scrollHeight;
    // 检测是否正在等待登录
    const isWaiting = data.job.status === 'running' &&
      logs.some(l => l.includes('[WAITING_FOR_LOGIN]')) &&
      !logs.some(l => l.includes('收到登录完成信号') || l.includes('秒超时，继续执行'));
    const btn = $('#resumeLoginBtn');
    if (btn) btn.classList.toggle('hidden', !isWaiting);
  } catch { state.selectedJobId = null; }
}

function renderJobList(jobs) {
  const root = $('#jobList');
  if (!root) return;
  if (!jobs.length) { root.innerHTML = '<span class="hint" style="padding:8px 0;display:block">暂无任务</span>'; return; }
  root.innerHTML = jobs.map(job => {
    const dc = job.status === 'running' || job.status === 'stopping' ? 'running'
             : job.status === 'failed' ? 'err'
             : job.status === 'finished' ? 'ok' : '';
    const sel = state.selectedJobId === job.id ? ' selected' : '';
    const rc = job.returncode != null ? ' rc:' + job.returncode : '';
    return '<div class="job-item' + sel + '" data-id="' + job.id + '">' +
      '<span class="jdot ' + dc + '"></span>' +
      '<div class="job-meta"><strong>' + escHtml(job.title) + '</strong>' +
      '<span class="hint">' + job.status + rc + ' · ' + relTime(job.started_at) + '</span></div>' +
      '</div>';
  }).join('');
  root.querySelectorAll('.job-item').forEach(item => {
    item.addEventListener('click', () => {
      state.selectedJobId = item.dataset.id;
      refreshJobLog();
      $$('#jobList .job-item').forEach(i => i.classList.toggle('selected', i.dataset.id === state.selectedJobId));
    });
  });
}

// ── Summary ───────────────────────────────────────────────────────────
async function refreshSummary() {
  try { const data = await api('/api/summary'); renderSummary(data); }
  catch (e) { showToast('刷新失败：' + e.message); }
}

function renderSummary(data) {
  // 顶部栏
  $('#clock').textContent = '当前 ' + shortTime(data.now);
  const pp = $('#proxyPill');
  pp.textContent = data.proxy.reachable
    ? '主代理在线 ' + data.proxy.host + ':' + data.proxy.port
    : '主代理未连接 ' + data.proxy.host + ':' + data.proxy.port;
  pp.className = 'pill ' + (data.proxy.reachable ? 'ok' : 'warn');
  const rolePill = $('#machineRolePill');
  const machineRole = data.machine_role || 'unconfigured';
  rolePill.textContent = machineRole === 'master' ? 'MASTER 数据主机' : (machineRole === 'replica' ? 'REPLICA 副本机' : '角色未配置');
  rolePill.className = 'pill ' + (machineRole === 'master' ? 'master' : (machineRole === 'replica' ? 'ok' : 'warn'));

  // 警告
  const wb = $('#warnings');
  const pendingReqs = data.pending_requests_count || 0;
  const msgs = [];
  if (data.warnings && data.warnings.length) msgs.push(...data.warnings);
  if (pendingReqs > 0) msgs.push(`📥 有 ${pendingReqs} 条待审采集需求` + (data.is_operator ? '，请前往 http://127.0.0.1:8765 批准' : '，等待操作员处理'));
  if (msgs.length) {
    wb.textContent = msgs.join(' · ');
    wb.classList.remove('hidden');
  } else {
    wb.classList.add('hidden');
  }

  // 指标卡片
  set('#mUnique',    fmtNumber(data.records.unique));
  set('#mEvents',    fmtNumber(data.records.events) + ' 条写入记录');
  set('#mRate',      data.records.success_rate + '%');
  set('#mSuccess',   fmtNumber(data.records.success) + ' 成功 / ' + fmtNumber(data.records.failed) + ' 失败 / ' + fmtNumber(data.records.pending) + ' 待采');
  set('#mRejection', fmtNumber(data.business.rejection));
  set('#mFwxx',      fmtNumber(data.business.fwxx_pending) + ' 待补发文');
  set('#mDue',       fmtNumber(data.business.update_due));
  set('#mTracked',   fmtNumber(data.business.tracked_total) + ' 跟踪中');
  set('#mDynamic',   fmtNumber(data.lists.dynamic));
  set('#mSearch',    fmtNumber(data.lists.search) + ' 输入申请号');

  // 概览 Tab
  renderTrendChart(data.daily_counts || []);
  renderSystemHealth(data);
  renderActiveJobs(data.jobs || []);

  // 采集控制 Tab
  const total = data.lists.search;
  const collected = data.lists.search_collected ?? 0;
  const pct = total > 0 ? Math.min(100, Math.round(collected / total * 100)) : 0;
  setStyle('#collectProgBar', 'width', pct + '%');
  set('#collectProgressHint', '已采集 ' + fmtNumber(collected) + ' / 输入 ' + fmtNumber(total) + '（' + pct + '%）');

  // 策略管理 Tab
  set('#stratTracked', fmtNumber(data.business.tracked_total));
  set('#stratDue',     fmtNumber(data.business.update_due));
  renderGroups(data.update_groups || []);

  // 发文采集 Tab
  const rej = data.business.rejection;
  const fwxxC = data.business.fwxx_collected;
  const fwxxP = data.business.fwxx_pending;
  const fwxxPct = rej > 0 ? Math.round(fwxxC / rej * 100) : 0;
  updateRing(fwxxPct);
  set('#fwxxPct',       fwxxPct + '%');
  set('#fwxxRejection', fmtNumber(rej));
  set('#fwxxCollected', fmtNumber(fwxxC));
  set('#fwxxPending',   fmtNumber(fwxxP));
  renderFwxxPending(data.fwxx_pending_list || []);

  // 数据分析 Tab
  renderBarList('#statusCounts',    data.status_counts    || []);
  renderBarList('#applicantCounts', data.applicant_counts || []);
  renderRejectionCompanies(data.rejection_companies || []);
  renderRecent(data.recent || []);

  // 数据管理 Tab
  set('#retryHistoryFailed', fmtNumber(data.records.failed ?? 0) + ' 条');
  set('#retryCount', fmtNumber(data.lists.failed_retry ?? 0) + ' 条');
  const jinfo = data.files && data.files.jsonl;
  set('#jsonlSize', jinfo ? fmtBytes(jinfo.size) : '—');

  // 系统配置 Tab
  set('#sysProxy',    data.proxy.host + ':' + data.proxy.port + (data.proxy.reachable ? ' ● 在线' : ' ● 离线'));
  set('#sysJsonlSize', jinfo ? fmtBytes(jinfo.size) + ' · ' + relTime(jinfo.mtime) : '—');
  set('#sysDynamic',  fmtNumber(data.lists.dynamic) + ' 条');
  set('#sysRetry',    fmtNumber(data.lists.retry) + ' 条');

  if (!state.configLoaded) {
    const ct = $('#configText');
    if (ct) { ct.value = JSON.stringify(data.config || {}, null, 2); state.configLoaded = true; }
  }

  // 角色切换：首次收到 is_operator 后设置 body class
  if (!state.roleDetermined) {
    state.roleDetermined = true;
    if (!data.is_operator) {
      document.body.classList.add('viewer-mode');
    }
  }

  // 需求队列（操作员）
  if (data.is_operator) refreshRequestQueue();
}

function set(sel, val) { const el = $(sel); if (el) el.textContent = val; }
function setStyle(sel, prop, val) { const el = $(sel); if (el) el.style[prop] = val; }

// ── Renderers ─────────────────────────────────────────────────────────
function renderTrendChart(daily) {
  const svg = $('#trendSvg');
  const lbls = $('#trendLabels');
  if (!svg || !daily.length) return;
  const W = 360, H = 64, px = 6, py = 6;
  const max = Math.max(...daily.map(d => d.count), 1);
  const pts = daily.map((d, i) => {
    const x = px + (i / Math.max(daily.length - 1, 1)) * (W - 2 * px);
    const y = py + (1 - d.count / max) * (H - 2 * py);
    return [+x.toFixed(1), +y.toFixed(1)];
  });
  const pline = pts.map(p => p[0] + ',' + p[1]).join(' ');
  const areaD = 'M' + pts[0][0] + ',' + H + ' ' +
    pts.map(p => 'L' + p[0] + ',' + p[1]).join(' ') +
    ' L' + pts[pts.length - 1][0] + ',' + H + ' Z';
  svg.innerHTML =
    '<defs><linearGradient id="tg" x1="0" y1="0" x2="0" y2="1">' +
    '<stop offset="0%" stop-color="#147a63" stop-opacity="0.25"/>' +
    '<stop offset="100%" stop-color="#147a63" stop-opacity="0"/>' +
    '</linearGradient></defs>' +
    '<path d="' + areaD + '" fill="url(#tg)"/>' +
    '<polyline fill="none" stroke="#147a63" stroke-width="2" stroke-linejoin="round" points="' + pline + '"/>' +
    pts.map((p, i) =>
      '<circle cx="' + p[0] + '" cy="' + p[1] + '" r="3" fill="#147a63"/>' +
      '<text x="' + p[0] + '" y="' + (H - 2) + '" text-anchor="middle" fill="#66736d" font-size="9" font-family="sans-serif">' +
      fmtNumber(daily[i].count) + '</text>'
    ).join('');
  if (lbls) lbls.innerHTML = daily.map(d => '<span>' + escHtml(d.date) + '</span>').join('');
}

function renderSystemHealth(data) {
  const hp = $('#healthProxy');
  if (hp) hp.innerHTML = data.proxy.reachable
    ? '<span class="pill ok">在线</span>'
    : '<span class="pill warn">离线</span>';
  set('#healthLastCollect', data.recent && data.recent.length ? relTime(data.recent[0].timestamp) : '—');
  const jinfo = data.files && data.files.jsonl;
  set('#healthLogSize', jinfo ? fmtBytes(jinfo.size) : '—');
  const einfo = data.files && data.files.excel;
  set('#healthExcel', einfo && einfo.exists ? fmtBytes(einfo.size) : '未生成');
}

function renderActiveJobs(jobs) {
  const root = $('#activeJobs');
  if (!root) return;
  if (!jobs.length) { root.innerHTML = '<span class="hint">暂无运行中的任务</span>'; return; }
  root.innerHTML = jobs.map(job =>
    '<div class="active-job-item">' +
    '<span class="running-dot"></span>' +
    '<span>' + escHtml(job.title) + '</span>' +
    '<span class="hint" style="margin-left:auto">' + job.status + ' · ' + relTime(job.started_at) + '</span>' +
    '</div>'
  ).join('');
}

function renderGroups(groups) {
  const root = $('#strategyGroups');
  if (!root) return;
  if (!groups.length) { root.innerHTML = '<div class="hint">暂无策略配置，请先运行"生成清单"</div>'; return; }
  const maxDue = Math.max(...groups.map(g => g.due), 1);
  root.innerHTML = groups.map(g => {
    const w = Math.max(3, Math.round(g.due / maxDue * 100));
    return '<div class="group-item">' +
      '<strong style="font-size:13px">' + g.frequency_days + ' 天</strong>' +
      '<div><div class="hint" style="margin-bottom:3px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">' +
      escHtml(g.statuses.join('、')) + '</div>' +
      '<div class="bar"><span style="width:' + w + '%"></span></div>' +
      '<div class="hint">' + fmtNumber(g.total) + ' 件</div></div>' +
      '<div><strong>' + fmtNumber(g.due) + '</strong><div class="hint">应检查</div></div>' +
      '<button class="btn secondary btn-run-freq" data-freq="' + g.frequency_days +
      '" style="font-size:12px;min-height:28px;padding:0 9px">采集</button>' +
      '</div>';
  }).join('');
}

function updateRing(pct) {
  const fg = $('#fwxxRingFg');
  if (!fg) return;
  fg.style.strokeDashoffset = 251.3 * (1 - pct / 100);
}

function renderFwxxPending(items) {
  const tbody = $('#fwxxPendingRows');
  if (!tbody) return;
  if (!items.length) { tbody.innerHTML = '<tr><td colspan="3" style="color:var(--muted)">暂无待补采数据</td></tr>'; return; }
  tbody.innerHTML = items.map(item =>
    '<tr><td>' + escHtml(item.application_no || '—') + '</td>' +
    '<td>' + escHtml(item.anjianywzt || '—') + '</td>' +
    '<td>' + shortTime(item.timestamp) + '</td></tr>'
  ).join('');
  set('#fwxxPendingHint', '待补 ' + items.length + ' 条（最新）');
}

function renderBarList(sel, rows) {
  const root = $(sel);
  if (!root) return;
  if (!rows.length) { root.innerHTML = '<span class="hint">暂无数据</span>'; return; }
  const max = Math.max(...rows.map(r => r[1]), 1);
  root.innerHTML = rows.map(([name, cnt]) =>
    '<div class="bar-row">' +
    '<span class="name" title="' + escHtml(name) + '">' + escHtml(name) + '</span>' +
    '<div class="bar-thin"><span style="width:' + Math.round(cnt / max * 100) + '%"></span></div>' +
    '<span class="cnt">' + fmtNumber(cnt) + '</span></div>'
  ).join('');
}

// ── 驳回企业列表 ──────────────────────────────────────────────────────────
let _rejCompanies = [];  // 全量缓存，用于过滤

function renderRejectionCompanies(companies) {
  _rejCompanies = companies;
  _applyRejCompanyFilter();
}

function _applyRejCompanyFilter() {
  const keyword = ($('#rejCompanySearch') || {}).value || '';
  const kw = keyword.trim().toLowerCase();
  const rows = kw
    ? _rejCompanies.filter(c => c.name.toLowerCase().includes(kw))
    : _rejCompanies;
  const tbody = $('#rejCompanyRows');
  if (!tbody) return;
  if (!rows.length) {
    tbody.innerHTML = '<tr><td colspan="4" class="hint" style="padding:8px">无匹配企业</td></tr>';
    return;
  }
  tbody.innerHTML = rows.map((c, idx) => {
    const realVal = c.real_total != null ? c.real_total : '';
    return '<tr>' +
      '<td style="max-width:280px"><div class="clip" title="' + escHtml(c.name) + '">' + escHtml(c.name) + '</div></td>' +
      '<td style="text-align:right">' + fmtNumber(c.invention_count) + '</td>' +
      '<td style="text-align:right"><input type="number" min="0" style="width:90px;text-align:right;padding:2px 4px" ' +
        'data-name="' + escHtml(c.name) + '" value="' + escHtml(String(realVal)) + '" placeholder="—"></td>' +
      '<td style="text-align:center"><button class="btn secondary" style="padding:2px 8px;font-size:12px" ' +
        'onclick="saveCompanyMeta(this)">保存</button></td>' +
      '</tr>';
  }).join('');
}

async function saveCompanyMeta(btn) {
  const row = btn.closest('tr');
  const input = row.querySelector('input[type=number]');
  const name = input.dataset.name;
  const val = input.value.trim();
  const real_total = val === '' ? null : parseInt(val, 10);
  if (val !== '' && isNaN(real_total)) { showToast('请输入有效数字'); return; }
  try {
    await api('/api/company-meta', { method: 'POST', body: JSON.stringify({ name, real_total }) });
    showToast('已保存：' + name);
    // 更新本地缓存
    const idx = _rejCompanies.findIndex(c => c.name === name);
    if (idx >= 0) _rejCompanies[idx].real_total = real_total;
  } catch(e) {
    showToast('保存失败：' + e.message);
  }
}

// 搜索框实时过滤
document.addEventListener('DOMContentLoaded', function() {
  const search = $('#rejCompanySearch');
  if (search) search.addEventListener('input', _applyRejCompanyFilter);
});

function downloadCompanyTemplate() {
  window.location.href = '/api/company-meta/template';
}

async function uploadCompanyMeta(input) {
  const file = input.files[0];
  if (!file) return;
  input.value = '';  // 允许重复上传同名文件
  const fd = new FormData();
  fd.append('file', file);
  try {
    showToast('上传中…');
    const res = await fetch('/api/company-meta/import', { method: 'POST', headers: writeHeaders(), body: fd });
    const d = await res.json();
    if (!res.ok) { showToast('上传失败：' + (d.error || res.status)); return; }
    showToast('导入完成：更新 ' + d.updated + ' 条，跳过 ' + d.skipped + ' 条');
    refreshSummary();
  } catch(e) {
    showToast('上传失败：' + e.message);
  }
}

function renderRecent(rows) {
  const tbody = $('#recentRows');
  if (!tbody) return;
  if (!rows.length) { tbody.innerHTML = '<tr><td colspan="7">暂无记录</td></tr>'; return; }
  tbody.innerHTML = rows.map(row => {
    const ok = row.status_code === 200;
    return '<tr>' +
      '<td>' + escHtml(row.application_no || '—') + '</td>' +
      '<td><span class="pill ' + (ok ? 'ok' : 'warn') + '">' + escHtml(String(row.status_code ?? '—')) + '</span></td>' +
      '<td>' + escHtml(row.anjianywzt || '—') + '</td>' +
      '<td><div class="clip" title="' + escHtml(row.zhuanlimc || '') + '">' + escHtml(row.zhuanlimc || '—') + '</div></td>' +
      '<td><div class="clip" title="' + escHtml(row.shenqingrxm || '') + '">' + escHtml(row.shenqingrxm || '—') + '</div></td>' +
      '<td>' + escHtml(String(row.response_time_ms ?? '—')) + '</td>' +
      '<td>' + shortTime(row.timestamp) + '</td></tr>';
  }).join('');
}

// ── Search List ───────────────────────────────────────────────────────
async function loadSearchList() {
  const data = await api('/api/search-list');
  const sl = $('#searchList');
  if (sl) sl.value = data.text || '';
  state.searchLoaded = true;
}

// ── Event Binding ────────────────────────────────────────────────────
function bindEvents() {
  // data-action 全局代理
  $$('[data-action]').forEach(btn => {
    btn.addEventListener('click', () => startJob(btn.dataset.action));
  });

  // 提交需求按钮
  const submitBtn = $('#submitReqBtn');
  if (submitBtn) submitBtn.addEventListener('click', submitRequest);

  // 凭证保存
  const saveCreds = $('#saveCreds');
  if (saveCreds) saveCreds.addEventListener('click', saveCredentials);

  // 增量导出/导入
  const exportBtn = $('#exportDeltaBtn');
  if (exportBtn) exportBtn.addEventListener('click', exportDelta);
  const importBtn = $('#importDeltaBtn');
  if (importBtn) importBtn.addEventListener('click', importDelta);

  // 策略分组"采集"按钮（动态生成，使用委托）
  document.getElementById('strategyGroups').addEventListener('click', e => {
    const btn = e.target.closest('.btn-run-freq');
    if (btn) startJob('strategy_collect', { frequency: btn.dataset.freq });
  });

  $('#runTest').addEventListener('click', () =>
    startJob('main_test', { count: $('#testCount').value }));

  $('#collectFwxxTest').addEventListener('click', () =>
    startJob('collect_fwxx', { count: $('#testCount').value }));

  $('#runUpdate').addEventListener('click', () =>
    startJob('main_update_dynamic', { file: $('#updateFile').value, count: $('#updateLimit').value }));

  // 立即用失败清单重新采集（复用 main_update_dynamic，预选 retry_failed.txt）
  const retryRecollectBtn = $('#retryRecollectBtn');
  if (retryRecollectBtn) retryRecollectBtn.addEventListener('click', () =>
    startJob('main_update_dynamic', { file: 'data/retry_failed.txt' }));

  // 筛选导出
  const retryBatchBtn = $('#retryBatchBtn');
  if (retryBatchBtn) retryBatchBtn.addEventListener('click', () =>
    startJob('retry_failed_batch', {
      batch_size: $('#retryBatchSize').value,
      batch_file: 'data/retry_batch_001.txt',
    }));

  const runRetryBatchBtn = $('#runRetryBatchBtn');
  if (runRetryBatchBtn) runRetryBatchBtn.addEventListener('click', () =>
    startJob('retry_failed_run_batch', {
      batch_file: 'data/retry_batch_001.txt',
      timeout: $('#retryTimeout').value,
    }));

  const exportApplicantSearch = $('#exportApplicantSearch');
  if (exportApplicantSearch) exportApplicantSearch.addEventListener('input', (e) =>
    renderApplicantCheckboxes(e.target.value));
  const exportPreviewBtn = $('#exportPreviewBtn');
  if (exportPreviewBtn) exportPreviewBtn.addEventListener('click', previewExport);
  const exportFilteredBtn = $('#exportFilteredBtn');
  if (exportFilteredBtn) exportFilteredBtn.addEventListener('click', exportFiltered);

  $('#generateStrategy').addEventListener('click', () =>
    startJob('strategy_generate', { frequency: $('#strategyFrequency').value }));

  $('#statusStrategy').addEventListener('click', () =>
    startJob('strategy_status', { frequency: $('#strategyFrequency').value }));

  $('#checkApp').addEventListener('click', () =>
    startJob('strategy_check', { app_no: $('#singleAppNo').value }));

  $('#autoPaginate').addEventListener('click', () =>
    startJob('public_auto_paginate', { delay: $('#pageDelay').value, max_pages: $('#maxPages').value }));

  $('#signalQueryReady').addEventListener('click', async () => {
    await api('/api/signal-query-ready', { method: 'POST', body: '{}' });
    showToast('已通知自动翻页脚本开始翻页');
  });

  $('#fwxxTestBtn').addEventListener('click', () =>
    startJob('collect_fwxx', { count: 5 }));

  $('#fwxxSingleBtn').addEventListener('click', () =>
    startJob('collect_fwxx_app', { app_no: $('#fwxxAppNo').value }));

  $('#stopJob').addEventListener('click', async () => {
    if (!state.selectedJobId) return;
    await api('/api/jobs/' + state.selectedJobId + '/stop', { method: 'POST', body: '{}' });
    showToast('已请求停止任务');
    await refreshJobs();
  });

  $('#saveSearchList').addEventListener('click', async () => {
    const sl = $('#searchList');
    if (!sl) return;
    await api('/api/search-list', { method: 'POST', body: JSON.stringify({ text: sl.value }) });
    showToast('申请号列表已保存');
    await loadSearchList();
    await refreshSummary();
  });

  $('#saveConfig').addEventListener('click', async () => {
    const ct = $('#configText');
    if (!ct) return;
    try {
      await api('/api/config', { method: 'POST', body: JSON.stringify({ text: ct.value }) });
      showToast('坐标配置已保存');
      state.configLoaded = false;
      await refreshSummary();
    } catch (e) { showToast('保存失败：' + e.message); }
  });

  $('#saveApiToken').addEventListener('click', () => {
    const input = $('#apiTokenInput');
    state.apiToken = input ? input.value.trim() : '';
    if (state.apiToken) localStorage.setItem('cnipaApiToken', state.apiToken);
    else localStorage.removeItem('cnipaApiToken');
    showToast(state.apiToken ? 'API token 已保存到此浏览器' : 'API token 已清除');
  });

  $('#resetConfig').addEventListener('click', async () => {
    await api('/api/config/reset', { method: 'POST', body: '{}' });
    showToast('旧坐标已备份，下次采集会重新记录');
    state.configLoaded = false;
    await refreshSummary();
  });

  $('#resumeLoginBtn').addEventListener('click', async () => {
    try {
      await api('/api/login-ready', { method: 'POST', body: '{}' });
      showToast('已发送登录完成信号，采集继续...');
      $('#resumeLoginBtn').classList.add('hidden');
    } catch (e) { showToast('发送失败：' + e.message); }
  });

  $('#loginDoneBtn').addEventListener('click', async () => {
    try {
      await api('/api/login-ready', { method: 'POST', body: '{}' });
      showToast('已发送登录完成信号，采集继续...');
      $('#loginBanner').classList.add('hidden');
    } catch (e) { showToast('发送失败：' + e.message); }
  });

  $('#updateNowBtn').addEventListener('click', () => {
    // 根据检查结果选择更新通道：git 模式走 upgrade_code，http 模式走 fetch_update
    const action = (lastUpdateCheck && lastUpdateCheck.method === 'http')
      ? 'fetch_update' : 'upgrade_code';
    $('#updateBanner').classList.add('hidden');
    startJob(action, {});  // startJob 内部已切到任务日志 Tab
  });

  $('#updateDismissBtn').addEventListener('click', () => {
    $('#updateBanner').classList.add('hidden');
  });

  $('#checkUpdateBtn').addEventListener('click', () => checkUpdate(true));

  $('#importAgencyBtn').addEventListener('click', async () => {
    const fi = $('#agencyFileInput');
    const hint = $('#agencyImportHint');
    if (!fi || !fi.files || !fi.files[0]) { if (hint) hint.textContent = '请先选择文件'; return; }
    if (hint) hint.textContent = '正在上传...';
    const formData = new FormData();
    formData.append('file', fi.files[0]);
    try {
      const res = await fetch('/api/import/agency', { method: 'POST', headers: writeHeaders(), body: formData });
      const d = await res.json();
      if (!res.ok) throw new Error(d.error || res.statusText);
      if (hint) hint.textContent =
        `✓ 更新 ${d.updated} 条` +
        (d.skipped_missing > 0 ? `，${d.skipped_missing} 条申请号不在库中` : '') +
        (d.skipped_no_app  > 0 ? `，${d.skipped_no_app} 条字段为空` : '') +
        (d.bad_rows        > 0 ? `，${d.bad_rows} 行格式错误` : '');
      showToast('代理机构导入完成：' + d.updated + ' 条');
    } catch (e) { if (hint) hint.textContent = '导入失败：' + e.message; }
  });
}

// ── 需求队列 ──────────────────────────────────────────────────────────
async function refreshRequestQueue() {
  try {
    const data = await api('/api/requests');
    renderRequestQueue(data.requests || []);
  } catch (_) {}
}

function renderRequestQueue(reqs) {
  const root = $('#reqQueueList');
  const hint = $('#reqQueueHint');
  if (!root) return;
  const pending = reqs.filter(r => r.status === 'pending');
  if (hint) hint.textContent = pending.length ? pending.length + ' 条待处理' : '暂无';
  if (!reqs.length) { root.innerHTML = '<span class="hint">暂无需求</span>'; return; }
  root.innerHTML = reqs.map(r => {
    const nos = (r.payload || []).join(', ');
    const badgeClass = {pending:'warn',executing:'ok',done:'ok',failed:'warn',rejected:'muted'}[r.status]||'muted';
    const btns = r.status === 'pending'
      ? '<button class="btn primary"   style="font-size:11px;min-height:24px;padding:0 8px" onclick="approveReq(\'' + escHtml(r.id) + '\')">批准</button>' +
        '<button class="btn secondary" style="font-size:11px;min-height:24px;padding:0 8px;margin-left:4px" onclick="rejectReq(\'' + escHtml(r.id) + '\')">拒绝</button>'
      : '';
    return '<div class="info-row" style="align-items:flex-start;gap:8px;padding:6px 0;border-bottom:1px solid var(--line)">' +
      '<span class="pill ' + badgeClass + '" style="flex-shrink:0">' + escHtml(r.status) + '</span>' +
      '<div style="flex:1;min-width:0"><div style="font-size:12px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="' + escHtml(nos) + '">' +
        escHtml((r.payload||[]).length + ' 个申请号') + (r.note ? ' · ' + escHtml(r.note) : '') +
      '</div><div class="hint" style="font-size:11px">' + escHtml(r.requester||'') + ' · ' + escHtml((r.created_at||'').substring(0,16)) + '</div></div>' +
      '<div style="flex-shrink:0">' + btns + '</div></div>';
  }).join('');
}

async function approveReq(id) {
  try {
    const data = await api('/api/requests/' + id + '/approve', { method: 'POST', body: '{}' });
    const parts = [];
    if (data.search_added > 0) parts.push(data.search_added + ' 个新申请号加入采集清单');
    if (data.already_in_db > 0) parts.push(data.already_in_db + ' 个已存在记录（策略系统会自动跟进）');
    showToast('已批准：' + (parts.length ? parts.join('，') : '无新增'));
    await refreshRequestQueue();
  } catch (e) { showToast('操作失败：' + e.message); }
}

async function rejectReq(id) {
  try {
    await api('/api/requests/' + id + '/reject', { method: 'POST', body: '{}' });
    showToast('已拒绝');
    await refreshRequestQueue();
  } catch (e) { showToast('操作失败：' + e.message); }
}

// ── 提交需求表单 ────────────────────────────────────────────────────
async function submitRequest() {
  const ta = $('#reqAppNos');
  const noteEl = $('#reqNote');
  const resultEl = $('#reqSubmitResult');
  const raw = (ta ? ta.value : '').trim();
  if (!raw) { if (resultEl) resultEl.textContent = '请填写至少一个申请号'; return; }
  // 过滤：只保留含数字的条目（排除"申请号"等纯文字表头）
  const app_nos = raw.split(/[\n,；;]+/)
    .map(s => s.trim())
    .filter(s => s && /\d/.test(s));
  try {
    const data = await api('/api/requests', {
      method: 'POST',
      body: JSON.stringify({ app_nos, note: noteEl ? noteEl.value : '' })
    });
    if (data.ok) {
      if (ta) ta.value = '';
      if (noteEl) noteEl.value = '';
      let msg = '✓ 已提交 ' + data.accepted + ' 个申请号，等待操作员审批';
      if (data.filtered > 0) msg += `（已过滤 ${data.filtered} 个无效条目）`;
      if (resultEl) resultEl.textContent = msg;
    } else {
      if (resultEl) resultEl.textContent = '⚠ ' + (data.error || '提交失败');
    }
  } catch (e) { if (resultEl) resultEl.textContent = '提交失败：' + e.message; }
}

// ── 登录凭证 ──────────────────────────────────────────────────────────
async function loadCredentials() {
  try {
    const d = await api('/api/credentials');
    const u = $('#credsUser'); if (u) u.value = d.username || '';
    const s = $('#credsStatus');
    if (s) s.textContent = d.password_set ? '✓ 密码已设置' : '⚠ 密码未设置';
  } catch (_) {}
}

async function saveCredentials() {
  const u = $('#credsUser'); const p = $('#credsPass'); const s = $('#credsStatus');
  try {
    await api('/api/credentials', { method: 'POST', body: JSON.stringify({
      username: u ? u.value : '',
      password: p ? p.value : ''
    })});
    if (p) p.value = '';
    if (s) s.textContent = '✓ 已保存';
    showToast('凭证已更新');
  } catch (e) { showToast('保存失败：' + e.message); }
}

// ── 增量数据互通 ───────────────────────────────────────────────────────
function exportDelta() {
  const inp = $('#deltaFrom');
  const hint = $('#deltaExportHint');
  if (!inp || !inp.value) { if (hint) hint.textContent = '请选择起始时间'; return; }
  const since = new Date(inp.value).toISOString();
  const url = '/api/export/delta?since=' + encodeURIComponent(since);
  if (hint) hint.textContent = '正在准备下载...';
  const a = document.createElement('a');
  a.href = url;
  a.download = 'delta_' + inp.value.slice(0,10) + '.jsonl';
  a.click();
  if (hint) hint.textContent = '已触发下载';
}

async function importDelta() {
  const fi = $('#deltaFileInput');
  const hint = $('#deltaImportHint');
  if (!fi || !fi.files || !fi.files[0]) { if (hint) hint.textContent = '请先选择文件'; return; }
  if (hint) hint.textContent = '正在导入...';
  const text = await fi.files[0].text();
  try {
    let res = await fetch('/api/import/delta', {
      method: 'POST',
      headers: writeHeaders({ 'Content-Type': 'application/x-ndjson' }),
      body: text
    });
    let d = await res.json();
    if (res.status === 409 && d.confirmation_required) {
      const s = d.summary;
      const confirmed = window.confirm(`MASTER 增量合并确认\n\n输入 ${s.records} 条\n新增申请号 ${s.new_applications} 条\n更新已有 ${s.updated_applications} 条\n时间范围 ${s.timestamp_from || '无'} → ${s.timestamp_to || '无'}\n\n确认继续？`);
      if (!confirmed) { if (hint) hint.textContent = '已取消导入'; return; }
      res = await fetch('/api/import/delta', {
        method: 'POST',
        headers: writeHeaders({ 'Content-Type': 'application/x-ndjson', 'X-CNIPA-Merge-Confirmed': 'yes' }),
        body: text
      });
      d = await res.json();
    }
    if (!res.ok) throw new Error(d.error || res.statusText);
    if (hint) hint.textContent = `✓ 已导入 ${d.imported} 条` + (d.bad_lines > 0 ? `，${d.bad_lines} 行格式错误` : '');
    showToast('导入完成：' + d.imported + ' 条');
  } catch (e) { if (hint) hint.textContent = '导入失败：' + e.message; }
}

// ── 更新检查 ──────────────────────────────────────────────────────────
// checkResult 保存最近一次检查结果，供「立即更新」决定走 git 还是 http
let lastUpdateCheck = null;

async function checkUpdate(showNoUpdateToast = false) {
  try {
    const d = await api('/api/check-update');
    lastUpdateCheck = d;
    const banner = $('#updateBanner');
    const text = $('#updateBannerText');
    if (d.has_update) {
      if (text) {
        text.textContent = d.method === 'git'
          ? `🆕 发现新版本：${d.pending_commits.length} 个新提交待更新`
          : `🆕 发现新版本：${d.local_version} → ${d.remote_version}`;
      }
      if (banner) banner.classList.remove('hidden');
    } else {
      if (banner) banner.classList.add('hidden');
      if (showNoUpdateToast) {
        showToast(d.error ? ('检查失败：' + d.error) : `已是最新版本（${d.local_version}）`);
      }
    }
    return d;
  } catch (e) {
    if (showNoUpdateToast) showToast('检查更新失败：' + e.message);
    return null;
  }
}

// ── 筛选导出 ──────────────────────────────────────────────────────────
let allApplicants = [];  // [{name, count}, ...]

async function loadApplicants() {
  const root = $('#exportApplicantList');
  if (!root) return;
  try {
    const d = await api('/api/applicants');
    allApplicants = d.applicants || [];
    renderApplicantCheckboxes('');
  } catch (e) {
    root.innerHTML = '<span class="hint">加载失败：' + escHtml(e.message) + '</span>';
  }
}

function renderApplicantCheckboxes(filter) {
  const root = $('#exportApplicantList');
  if (!root) return;
  const kw = (filter || '').trim().toLowerCase();
  // 记住已勾选的，过滤后保持选中状态
  const checked = new Set(
    Array.from(root.querySelectorAll('input:checked')).map(c => c.value)
  );
  const list = kw
    ? allApplicants.filter(a => a.name.toLowerCase().includes(kw))
    : allApplicants;
  if (!list.length) { root.innerHTML = '<span class="hint">无匹配申请人</span>'; return; }
  root.innerHTML = list.slice(0, 300).map(a =>
    '<label style="display:block;padding:2px 0;cursor:pointer">' +
    '<input type="checkbox" value="' + escHtml(a.name) + '"' +
      (checked.has(a.name) ? ' checked' : '') + '> ' +
    escHtml(a.name) + ' <span class="hint">(' + a.count + ')</span></label>'
  ).join('') + (list.length > 300 ? '<div class="hint">仅显示前 300 项，请用关键词缩小范围</div>' : '');
}

function collectExportFilters() {
  const applicants = Array.from(
    $('#exportApplicantList').querySelectorAll('input:checked')
  ).map(c => c.value);
  // date input 是 YYYY-MM-DD；采集时间转 ISO（含 Z），驳回日期保持 YYYY-MM-DD
  const tsFrom = $('#exportTsFrom').value;
  const tsTo = $('#exportTsTo').value;
  return {
    applicants,
    timestamp_from: tsFrom ? tsFrom + 'T00:00:00Z' : '',
    timestamp_to:   tsTo ? tsTo + 'T23:59:59Z' : '',
    rejection_from: $('#exportRejFrom').value || '',
    rejection_to:   $('#exportRejTo').value || '',
  };
}

async function previewExport() {
  const hint = $('#exportFilterHint');
  if (hint) hint.textContent = '统计中...';
  try {
    const res = await fetch('/api/export/excel-filtered?preview=true', {
      method: 'POST',
      headers: writeHeaders({ 'Content-Type': 'application/json' }),
      body: JSON.stringify(collectExportFilters()),
    });
    const d = await res.json();
    if (hint) hint.textContent = '符合条件：' + d.count + ' 条';
  } catch (e) {
    if (hint) hint.textContent = '统计失败：' + e.message;
  }
}

async function exportFiltered() {
  const hint = $('#exportFilterHint');
  if (hint) hint.textContent = '正在生成 Excel...';
  try {
    const res = await fetch('/api/export/excel-filtered', {
      method: 'POST',
      headers: writeHeaders({ 'Content-Type': 'application/json' }),
      body: JSON.stringify(collectExportFilters()),
    });
    if (!res.ok) {
      const err = await res.json().catch(() => ({}));
      throw new Error(err.error || res.statusText);
    }
    const blob = await res.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    const cd = res.headers.get('Content-Disposition') || '';
    const m = cd.match(/filename="([^"]+)"/);
    a.download = m ? m[1] : 'patents_filtered.xlsx';
    a.click();
    URL.revokeObjectURL(url);
    if (hint) hint.textContent = '✓ 已导出';
  } catch (e) {
    if (hint) hint.textContent = '导出失败：' + e.message;
  }
}

// ── Boot ─────────────────────────────────────────────────────────────
async function boot() {
  initTabRouting();
  bindEvents();
  await loadOperatorToken();
  await Promise.all([refreshSummary(), refreshJobs(), loadSearchList(), loadCredentials()]);
  setInterval(refreshSummary, 5000);
  setInterval(refreshJobs, 2500);
  checkUpdate();                              // 启动时检查一次
  setInterval(checkUpdate, 3600000);          // 每小时检查一次
  loadApplicants();                           // 加载筛选导出的申请人列表
}

boot().catch(e => showToast(e.message));
"""


# ══════════════════════════════════════════════════════════════════════
#  HTTP 服务器
# ══════════════════════════════════════════════════════════════════════

class DashboardHandler(BaseHTTPRequestHandler):
    server_version = SERVER_VERSION
    job_manager: JobManager

    def log_message(self, format: str, *args: Any) -> None:
        timestamp = datetime.now().strftime("%H:%M:%S")
        print(f"[{timestamp}] {self.address_string()} {format % args}")

    @property
    def is_operator(self) -> bool:
        return self.client_address[0] == '127.0.0.1'

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        try:
            if path == "/":
                self.send_text(HTML, "text/html; charset=utf-8")
            elif path == "/app.css":
                self.send_text(CSS, "text/css; charset=utf-8")
            elif path == "/app.js":
                self.send_text(JS, "application/javascript; charset=utf-8")
            elif path == "/api/summary":
                summary = build_summary(self.job_manager)
                summary['is_operator'] = self.is_operator
                summary['machine_role'] = read_machine_role()
                self.send_json(summary)
            elif path == "/api/alert-status":
                self.send_json(read_alert_status())
            elif path == "/api/operator-token":
                if not self.is_operator:
                    self.send_json({"error": "仅本机可自动读取 API token"}, status=403)
                    return
                self.send_json({"token": ensure_api_token()})
            elif path == "/api/check-update":
                # 同步调 check_update.py，解析最后一行 JSON 返回前端（轻量，不进 JobManager）
                proc = subprocess.run(
                    [resolve_task_python(), "-u", "check_update.py"],
                    cwd=str(BASE_DIR), capture_output=True, text=True,
                    encoding="utf-8", errors="replace", timeout=40,
                )
                result = None
                for line in reversed(proc.stdout.splitlines()):
                    line = line.strip()
                    if line.startswith("{"):
                        try:
                            result = json.loads(line)
                            break
                        except json.JSONDecodeError:
                            continue
                if result is None:
                    self.send_json({"has_update": False, "error": "检查脚本无有效输出"}, status=502)
                else:
                    self.send_json(result)
            elif path == "/api/applicants":
                # 全部不同申请人（2000+），供筛选导出下拉；按数量降序
                applicants = [
                    {"name": name, "count": count}
                    for name, count in _patents_db.list_applicants()
                ]
                self.send_json({"applicants": applicants})
            elif path == "/api/requests":
                if not self.is_operator:
                    self.send_json({"error": "仅操作员可查看需求列表"}, status=403)
                    return
                reqs = _patents_db.list_requests()
                for r in reqs:
                    if r.get('status') == 'executing' and r.get('job_id'):
                        job = self.job_manager.get_job(r['job_id'])
                        if job:
                            _patents_db.sync_request_status(r['id'], job.status, job.returncode)
                            r['status'] = job.status if job.status != 'finished' else (
                                'done' if job.returncode == 0 else 'failed'
                            )
                self.send_json({"requests": reqs})
            elif path == "/api/jobs":
                self.send_json({"jobs": self.job_manager.list_jobs()})
            elif path.startswith("/api/jobs/"):
                self.handle_get_job(path)
            elif path == "/api/search-list":
                self.send_json({"text": safe_read_text(SEARCH_LIST_FILE), "path": str(SEARCH_LIST_FILE)})
            elif path == "/api/config":
                self.send_json({"text": safe_read_text(CONFIG_FILE, "{}"), "path": str(CONFIG_FILE)})
            elif path == "/api/credentials":
                if not self.is_operator:
                    self.send_json({"error": "仅操作员可查看凭证"}, status=403)
                    return
                pairs = _parse_env_file(BASE_DIR / ".env")
                self.send_json({
                    "username": pairs.get("CNIPA_USERNAME", ""),
                    "password_set": bool(pairs.get("CNIPA_PASSWORD")),
                })
            elif path == "/api/company-meta/template":
                # 生成「跟踪企业 + 驳回企业」合并补录模板 Excel
                import tempfile
                try:
                    import openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment
                except ImportError:
                    self.send_json({"error": "openpyxl 未安装"}, status=500)
                    return
                # 读取跟踪状态列表
                focus = safe_json_load(DATA_DIR / "focus_strategy.json", {})
                tracked_statuses = list((focus.get("status_breakdown") or {}).keys())
                # 从 DB 一次性获取合并后的企业列表
                company_rows = _patents_db.get_company_meta_rows(tracked_statuses)
                # 合并现有 company_meta.json 的已补录数据
                meta = safe_json_load(COMPANY_META_FILE, {})
                if not isinstance(meta, dict):
                    meta = {}
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "企业实际专利数"
                headers = [
                    "企业名",
                    "跟踪中件数（在审专利）",
                    "库内发明专利总数（含其他状态）",
                    "实际专利总数（手动填写）",
                ]
                ws.append(headers)
                header_fill = PatternFill("solid", fgColor="4472C4")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                # 说明行
                note = "★ 填写说明：D 列填入该企业在国知局的发明专利申请总数，保存后上传至系统。A 列企业名请勿修改。"
                ws.append([note, "", "", ""])
                ws.cell(2, 1).font = Font(color="FF0000", italic=True)
                ws.merge_cells("A2:D2")
                # 数据行
                for c in company_rows:
                    real_total = (meta.get(c["name"]) or {}).get("real_total")
                    ws.append([
                        c["name"],
                        c["tracked_count"] or "",
                        c["total_count"] or "",
                        real_total if real_total is not None else "",
                    ])
                # 列宽
                ws.column_dimensions["A"].width = 52
                ws.column_dimensions["B"].width = 22
                ws.column_dimensions["C"].width = 26
                ws.column_dimensions["D"].width = 24
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                    wb.save(tmp.name)
                    tmp_path = tmp.name
                data = Path(tmp_path).read_bytes()
                os.unlink(tmp_path)
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", 'attachment; filename="company_meta_template.xlsx"')
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
            elif path.startswith("/api/export/delta"):
                qs = parse_qs(parsed.query)
                since = (qs.get("since") or [""])[0].strip()
                if not since:
                    self.send_json({"error": "缺少 since 参数（格式：2026-05-01T00:00:00Z）"}, status=400)
                    return
                try:
                    datetime.fromisoformat(since.replace('Z', '+00:00'))
                except ValueError:
                    self.send_json({"error": "since 不是有效的 ISO 时间戳"}, status=400)
                    return
                records = _patents_db.export_delta(since)
                lines = "\n".join(json.dumps(r, ensure_ascii=False) for r in records)
                data = lines.encode("utf-8")
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/x-ndjson; charset=utf-8")
                self.send_header("Content-Disposition", f'attachment; filename="delta_{since[:10]}.jsonl"')
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
            elif path.startswith("/download/"):
                self.handle_download(path)
            else:
                self.send_error(HTTPStatus.NOT_FOUND, "Not found")
        except Exception as exc:
            self.send_json({"error": str(exc)}, status=500)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        if path != "/api/requests" and not api_token_matches(self.headers.get('X-CNIPA-Token')):
            self.send_json({"error": "写操作需要有效的 X-CNIPA-Token"}, status=401)
            return
        try:
            if path == "/api/jobs":
                payload = self.read_json_body()
                job = self.job_manager.start(payload.get("action", ""), payload.get("params") or {})
                self.send_json({"job": job.to_dict(include_logs=True)}, status=201)
            elif path.startswith("/api/jobs/") and path.endswith("/stop"):
                job_id = _parse_path_segment(path, 3)
                ok = self.job_manager.stop(job_id)
                self.send_json({"ok": ok})
            elif path == "/api/company-meta":
                payload = self.read_json_body()
                name = str(payload.get("name", "")).strip()
                real_total = payload.get("real_total")
                if not name:
                    self.send_json({"error": "企业名不能为空"}, status=400)
                    return
                meta = safe_json_load(COMPANY_META_FILE, {})
                if not isinstance(meta, dict):
                    meta = {}
                entry = meta.setdefault(name, {})
                if real_total is None:
                    entry.pop("real_total", None)
                else:
                    entry["real_total"] = int(real_total)
                if not entry:
                    meta.pop(name, None)
                COMPANY_META_FILE.parent.mkdir(parents=True, exist_ok=True)
                _write_text_atomic(COMPANY_META_FILE, json.dumps(meta, ensure_ascii=False, indent=2) + "\n")
                self.send_json({"ok": True})
            elif path == "/api/company-meta/import":
                # 解析上传的 Excel，批量更新 company_meta.json
                import cgi
                try:
                    import openpyxl
                except ImportError:
                    self.send_json({"error": "openpyxl 未安装，无法解析 Excel"}, status=500)
                    return
                content_type = self.headers.get("Content-Type", "")
                length = int(self.headers.get("Content-Length", 0))
                raw_body = self.rfile.read(length)
                environ = {"REQUEST_METHOD": "POST", "CONTENT_TYPE": content_type, "CONTENT_LENGTH": str(length)}
                fs = cgi.FieldStorage(fp=io.BytesIO(raw_body), environ=environ, keep_blank_values=True)
                file_item = fs.getvalue("file")
                if file_item is None:
                    self.send_json({"error": "未收到文件"}, status=400)
                    return
                import tempfile
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                    tmp.write(file_item if isinstance(file_item, bytes) else file_item.encode())
                    tmp_path = tmp.name
                updated = skipped = 0
                try:
                    wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
                    ws = wb.active
                    meta = safe_json_load(COMPANY_META_FILE, {})
                    if not isinstance(meta, dict):
                        meta = {}
                    # 第1行为表头，第2行为说明，从第3行起是数据
                    # A列=企业名，D列(index 3)=实际专利总数（兼容旧3列格式：C列=index 2）
                    for row in ws.iter_rows(min_row=3, values_only=True):
                        name_val = row[0] if row else None
                        # 4列新格式：D列；3列旧格式：C列
                        total_val = row[3] if len(row) > 3 else (row[2] if len(row) > 2 else None)
                        if not name_val or str(name_val).strip() == "":
                            continue
                        name_str = str(name_val).strip()
                        if total_val is None or str(total_val).strip() == "":
                            skipped += 1
                            continue
                        try:
                            real_total = int(float(str(total_val)))
                        except (ValueError, TypeError):
                            skipped += 1
                            continue
                        meta.setdefault(name_str, {})["real_total"] = real_total
                        updated += 1
                    wb.close()
                    COMPANY_META_FILE.parent.mkdir(parents=True, exist_ok=True)
                    _write_text_atomic(COMPANY_META_FILE, json.dumps(meta, ensure_ascii=False, indent=2) + "\n")
                except Exception as exc:
                    self.send_json({"error": f"解析 Excel 失败：{exc}"}, status=400)
                    return
                finally:
                    os.unlink(tmp_path)
                self.send_json({"ok": True, "updated": updated, "skipped": skipped})
            elif path == "/api/search-list":
                payload = self.read_json_body()
                search_app_nos = parse_app_no_list(str(payload.get("text", "")))
                text = "\n".join(search_app_nos)
                if text:
                    text += "\n"
                SEARCH_LIST_FILE.parent.mkdir(parents=True, exist_ok=True)
                _write_text_atomic(SEARCH_LIST_FILE, text)
                self.send_json({"ok": True, "lines": len(search_app_nos)})
            elif path == "/api/config":
                payload = self.read_json_body()
                text = str(payload.get("text", "{}"))
                parsed_json = json.loads(text)
                CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)
                _write_text_atomic(CONFIG_FILE, json.dumps(parsed_json, ensure_ascii=False, indent=2) + "\n")
                self.send_json({"ok": True})
            elif path == "/api/config/reset":
                backup = None
                if CONFIG_FILE.exists():
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    backup = CONFIG_FILE.with_name(f"config_backup_{timestamp}.json")
                    CONFIG_FILE.rename(backup)
                self.send_json({"ok": True, "backup": str(backup) if backup else None})
            elif path == "/api/login-ready":
                flag = DATA_DIR / "login_ready.flag"
                flag.parent.mkdir(parents=True, exist_ok=True)
                flag.touch()
                self.send_json({"ok": True})
            elif path == "/api/signal-query-ready":
                # 向 auto_paginate.py 发出"查询已就绪"信号，与 login-ready 机制对称
                flag = DATA_DIR / "public_query_ready.flag"
                flag.parent.mkdir(parents=True, exist_ok=True)
                flag.touch()
                self.send_json({"ok": True})
            elif path == "/api/credentials":
                if not self.is_operator:
                    self.send_json({"error": "仅操作员可修改凭证"}, status=403)
                    return
                payload = self.read_json_body()
                username = str(payload.get("username", "")).strip()
                password = str(payload.get("password", "")).strip()
                env_file = BASE_DIR / ".env"
                pairs = _parse_env_file(env_file)
                if username:
                    pairs["CNIPA_USERNAME"] = username
                if password:
                    pairs["CNIPA_PASSWORD"] = password
                _write_text_atomic(env_file, "\n".join(f"{k}={v}" for k, v in pairs.items()) + "\n")
                self.send_json({"ok": True})
            elif path == "/api/export/excel-filtered":
                payload = self.read_json_body()
                applicants = payload.get("applicants") or None
                ts_from    = payload.get("timestamp_from") or None
                ts_to      = payload.get("timestamp_to") or None
                rej_from   = payload.get("rejection_from") or None
                rej_to     = payload.get("rejection_to") or None
                # 去空值：前端可能传 ""
                if applicants:
                    applicants = [a for a in applicants if a.strip()]
                    if not applicants:
                        applicants = None

                records = _patents_db.query_filtered(
                    applicants=applicants, ts_from=ts_from, ts_to=ts_to,
                    rejection_from=rej_from, rejection_to=rej_to,
                )

                # 预览模式：只返回数量
                qs = parse_qs(parsed.query)
                if qs.get("preview"):
                    self.send_json({"count": len(records)})
                    return

                # 生成 Excel 并返回文件下载
                import tempfile
                from detection_logger import DetectionLogger
                logger = DetectionLogger()
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                tmp.close()
                # 临时替换 _load_records 让 export_to_excel 使用筛选后的记录
                orig_load = logger._load_records
                logger._load_records = lambda: records
                try:
                    logger.export_to_excel(tmp.name)
                finally:
                    logger._load_records = orig_load

                excel_data = Path(tmp.name).read_bytes()
                os.unlink(tmp.name)

                ts_label = datetime.now().strftime("%Y%m%d_%H%M%S")
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="patents_filtered_{ts_label}.xlsx"')
                self.send_header("Content-Length", str(len(excel_data)))
                self.end_headers()
                self.wfile.write(excel_data)
            elif path == "/api/import/delta":
                if not self.is_operator:
                    self.send_json({"error": "仅操作员可导入数据"}, status=403)
                    return
                length = int(self.headers.get("Content-Length", 0))
                body = self.rfile.read(length).decode("utf-8")
                records, bad = [], 0
                for line in body.splitlines():
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        records.append(json.loads(line))
                    except json.JSONDecodeError:
                        bad += 1
                import_summary = _patents_db.summarize_record_import(records)
                if read_machine_role() == MASTER_ROLE and self.headers.get('X-CNIPA-Merge-Confirmed') != 'yes':
                    self.send_json({
                        "error": "master 增量合并需要确认",
                        "confirmation_required": True,
                        "summary": import_summary,
                    }, status=409)
                    return
                imported = _patents_db.upsert_batch(records)
                self.send_json({"ok": True, "imported": imported, "bad_lines": bad, "summary": import_summary})
            elif path == "/api/import/agency":
                if not self.is_operator:
                    self.send_json({"error": "仅操作员可导入数据"}, status=403)
                    return
                # 解析 multipart/form-data 中的文件
                import cgi
                content_type = self.headers.get("Content-Type", "")
                length = int(self.headers.get("Content-Length", 0))
                raw_body = self.rfile.read(length)
                environ = {
                    "REQUEST_METHOD": "POST",
                    "CONTENT_TYPE": content_type,
                    "CONTENT_LENGTH": str(length),
                }
                fs = cgi.FieldStorage(
                    fp=io.BytesIO(raw_body),
                    environ=environ,
                    keep_blank_values=True,
                )
                file_item = fs.getvalue("file")
                if file_item is None:
                    self.send_json({"error": "未收到文件"}, status=400)
                    return
                # 写临时文件，由 import_agency_csv 解析
                import tempfile
                import os as _os
                filename = fs["file"].filename if hasattr(fs["file"], "filename") else "upload.csv"
                suffix = Path(filename).suffix.lower() or ".csv"
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                    tmp.write(file_item if isinstance(file_item, bytes) else file_item.encode())
                    tmp_path = tmp.name
                try:
                    from import_agency_csv import import_agency
                    stats = import_agency(Path(tmp_path), dry_run=False)
                    self.send_json({"ok": True, **stats})
                except (ValueError, ImportError) as exc:
                    self.send_json({"error": str(exc)}, status=400)
                finally:
                    _os.unlink(tmp_path)
            elif path == "/api/requests":
                payload = self.read_json_body()
                raw_nos = [str(a).strip() for a in (payload.get("app_nos") or []) if str(a).strip()]
                # 只接受含数字的条目，排除"申请号"等纯文字表头
                app_nos = [a for a in raw_nos if any(c.isdigit() for c in a)]
                filtered = len(raw_nos) - len(app_nos)
                if not app_nos:
                    self.send_json({"error": "app_nos 不能为空（纯文字条目已过滤）"}, status=400)
                    return
                if len(app_nos) > MAX_REQUEST_APP_NOS:
                    self.send_json({"error": f"单次最多提交 {MAX_REQUEST_APP_NOS} 个申请号"}, status=400)
                    return
                note = str(payload.get("note", ""))[:MAX_NOTE_LEN]
                req_id = _patents_db.submit_request(app_nos, self.client_address[0], note)
                if not req_id:
                    self.send_json({"ok": False, "error": "申请号已在待处理队列中，请勿重复提交"}, status=409)
                    return
                self.send_json({"ok": True, "id": req_id, "filtered": filtered, "accepted": len(app_nos)}, status=201)
            elif path.startswith("/api/requests/") and path.endswith("/approve"):
                if not self.is_operator:
                    self.send_json({"error": "仅操作员可批准需求"}, status=403)
                    return
                req_id = _parse_path_segment(path, 3)
                reqs = _patents_db.list_requests()
                req = next((r for r in reqs if r['id'] == req_id), None)
                if not req or req['status'] != 'pending':
                    self.send_json({"error": "需求不存在或状态不是 pending"}, status=404)
                    return
                # 规范化申请号：只有不在 DB 中的才加入采集清单，已有记录由策略系统按节奏更新
                all_in_db = _patents_db.get_all_app_nos()
                to_search = []
                for raw in req['payload']:
                    norm = normalize_app_no(raw)
                    if norm and norm not in all_in_db:
                        to_search.append(norm)

                already_in_db = len(req['payload']) - len(to_search)

                def _atomic_append(file_path: Path, new_entries: list[str]) -> int:
                    existing_set = set()
                    if file_path.exists():
                        existing_set = {ln.strip() for ln in file_path.read_text(encoding='utf-8').splitlines() if ln.strip()}
                    fresh = [e for e in new_entries if e not in existing_set]
                    if fresh:
                        merged = sorted(existing_set) + fresh
                        tmp = file_path.with_suffix('.tmp')
                        tmp.write_text('\n'.join(merged) + '\n', encoding='utf-8')
                        tmp.replace(file_path)
                    return len(fresh)

                search_added = _atomic_append(SEARCH_LIST_FILE, to_search) if to_search else 0

                _patents_db.approve_request(req_id, None)
                _patents_db.sync_request_status(req_id, 'finished', 0)
                self.send_json({"ok": True, "search_added": search_added, "already_in_db": already_in_db})
            elif path.startswith("/api/requests/") and path.endswith("/reject"):
                if not self.is_operator:
                    self.send_json({"error": "仅操作员可拒绝需求"}, status=403)
                    return
                req_id = _parse_path_segment(path, 3)
                _patents_db.reject_request(req_id)
                self.send_json({"ok": True})
            else:
                self.send_error(HTTPStatus.NOT_FOUND, "Not found")
        except json.JSONDecodeError:
            self.send_json({"error": "JSON 格式不正确"}, status=400)
        except ValueError as exc:
            self.send_json({"error": str(exc)}, status=400)
        except Exception as exc:
            self.send_json({"error": str(exc)}, status=500)

    def handle_get_job(self, path: str) -> None:
        job_id = _parse_path_segment(path, 3)
        job = self.job_manager.get_job(job_id)
        if not job:
            self.send_json({"error": "任务不存在"}, status=404)
            return
        self.send_json({"job": job.to_dict(include_logs=True)})

    def handle_download(self, path: str) -> None:
        key = path.rsplit("/", 1)[-1]
        target = DOWNLOADS.get(key)
        if not target or not target.exists():
            self.send_error(HTTPStatus.NOT_FOUND, "File not found")
            return
        content_type = mimetypes.guess_type(str(target))[0] or "application/octet-stream"
        data = target.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Content-Disposition", f'attachment; filename="{target.name}"')
        self.end_headers()
        self.wfile.write(data)

    def read_json_body(self) -> dict[str, Any]:
        length = int(self.headers.get("Content-Length", "0") or "0")
        if length <= 0:
            return {}
        if length > MAX_BODY_BYTES:
            raise ValueError(f"请求体超过大小限制（最大 {MAX_BODY_BYTES // 1024} KB）")
        body = self.rfile.read(length).decode("utf-8", errors="replace")
        payload = json.loads(body)
        if not isinstance(payload, dict):
            raise ValueError("请求体必须是 JSON 对象")
        return payload

    def send_text(self, text: str, content_type: str, status: int = 200) -> None:
        data = text.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(data)

    def send_json(self, payload: dict[str, Any], status: int = 200) -> None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(data)


def run_server(host: str, port: int) -> None:
    ensure_api_token()
    job_manager = JobManager()
    DashboardHandler.job_manager = job_manager
    server = ThreadingHTTPServer((host, port), DashboardHandler)
    print(f"{APP_NAME} 已启动: http://{host}:{port}")
    print("按 Ctrl+C 停止控制台")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n正在停止控制台...")
    finally:
        server.server_close()


def main() -> None:
    parser = argparse.ArgumentParser(description="启动 CNIPA 本地可视化控制台")
    parser.add_argument("--host", default="0.0.0.0", help="监听地址，默认 0.0.0.0（局域网可访问）")
    parser.add_argument("--port", type=int, default=8765, help="监听端口，默认 8765")
    args = parser.parse_args()
    run_server(args.host, args.port)


if __name__ == "__main__":
    main()
